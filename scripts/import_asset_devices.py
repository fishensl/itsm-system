# -*- coding: utf-8 -*-
"""鄱阳湖水文资产登记表 → 设备管理 导入脚本。

解析《鄱阳湖水文网络设备与会议设备资产登记表.xlsx》「网络设备」sheet：
- 自动识别机柜组头行（序号列非数字，如「机房机柜1」「机房机柜1背面」）
- 列映射：设备名称/IP/品牌/型号/序列号/位置/建设时间/授权到期/系统版本/规则库版本/
  是否在用/硬件维修情况/接口类型及数量
- device_type 按设备名称关键词推断（防火墙/交换机/路由器/服务器/存储/录像机/光猫/KVM/
  无线AP/上网行为管理/日志审计/IPS/光传输/其他）
- 「用户名密码」列不导入（明文密码不进库，安全约定）
- 幂等：按 (客户, 设备名称) 查重，已存在默认跳过（--update 更新字段）
- 导入完成后刷新客户 device_count/等级（统一入口）

用法（项目根目录）：
    python scripts/import_asset_devices.py <xlsx路径>            # 预览（不写库）
    python scripts/import_asset_devices.py <xlsx路径> --apply    # 实际导入
    python scripts/import_asset_devices.py <xlsx路径> --apply --update
    python scripts/import_asset_devices.py <xlsx路径> --apply --create-customer
    python scripts/import_asset_devices.py <xlsx路径> --customer-id 123
"""
import os
import re
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402

from app import create_app  # noqa: E402
from models import db, Customer, Device, Region  # noqa: E402
from utils.json_fields import dumps_json  # noqa: E402

CUSTOMER_NAME = '鄱阳湖水文水资源监测中心'
GROUP_PATTERN = re.compile(r'机房机柜\d+(背面)?')

# 名称关键词 → device_type（顺序敏感：先精确再宽泛）
TYPE_RULES = [
    ('防火墙', '防火墙'),
    ('IPS', 'IPS'),
    ('入侵防御', 'IPS'),
    ('交换机', '交换机'),
    ('路由器', '路由器'),
    ('无线控制器', '无线AP'),
    ('上网行为', '上网行为管理'),
    ('行为管理', '上网行为管理'),
    ('日志审计', '日志审计'),
    ('录像机', '录像机'),
    ('存储', '存储'),
    ('服务器', '服务器'),
    ('ESXI', '服务器'),
    ('主机', '服务器'),
    ('光猫', '光猫'),
    ('光分', '光传输'),
    ('KVM', 'KVM'),
    ('机房卫士', '其他'),
    ('IP RAN', '路由器'),
    ('RAN', '路由器'),
]


def infer_device_type(name):
    """按设备名称关键词推断设备类型"""
    for kw, dtype in TYPE_RULES:
        if kw.lower() in name.lower():
            return dtype
    return '其他'


def group_of(location):
    """从位置字符串（如「机房机柜2·37U」）提取机柜分组名"""
    return location.split('·', 1)[0] if location else ''


def _clean(v):
    """清洗单元格：去首尾空白与内嵌换行"""
    if v is None:
        return ''
    return re.sub(r'\s+', '', str(v).strip())


def _to_date(v):
    """datetime 或 %Y-%m-%d 字符串 → date；失败返回 None"""
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    from services.device_service import _parse_date
    return _parse_date(_clean(v))


def parse_asset_excel(path):
    """解析资产表 → 设备字典列表（纯函数，不碰库）

    返回 (rows, groups_skipped, errors)；rows 元素字段与 Device 构造参数对齐。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['网络设备'] if '网络设备' in wb.sheetnames else wb.active
    # 表头行搜索：第 1 行通常是文档标题（"…资产登记表"），第 2 行才是列标题
    header_row_idx = None
    col = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        cols = {}
        for idx, cell in enumerate(row, start=1):
            v = _clean(cell)
            if v:
                cols[v] = idx
        if '设备名称' in cols:
            header_row_idx = r_idx
            col = cols
            break
    if '设备名称' not in col:
        raise ValueError('Excel 缺少必需列「设备名称」，请确认文件为资产登记表格式')

    rows, groups_skipped, errors = [], 0, []
    group = ''
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        seq = _clean(row[0]) if row else ''
        name = _clean(row[col['设备名称'] - 1]) if col['设备名称'] - 1 < len(row) else ''
        if not seq or not seq.isdigit():
            # 组头行（A 列如「机房机柜1」「机房机柜1背面」）：更新当前机柜分组
            cell0 = _clean(row[0]) if row else ''
            if GROUP_PATTERN.match(cell0) or cell0 == '网络设备':
                group = cell0
                groups_skipped += 1
            continue
        if not name:
            errors.append(f'{group} 第{seq}行：设备名称为空，跳过')
            continue

        def _c(key):
            i = col.get(key)
            return _clean(row[i - 1]) if i and i - 1 < len(row) else ''

        location = f'{group}·{_c("位置")}' if group else _c('位置')
        interface_raw = _c('接口类型及数量')
        interfaces = [p.strip() for p in re.split(r'[，,、;；\n/]', interface_raw) if p.strip()]
        rows.append({
            'device_name': name,
            'device_type': infer_device_type(name),
            'brand': _c('品牌'),
            'model': _c('型号'),
            'serial_number': _c('序列号'),
            'ip_address': _c('IP地址'),
            'location': location,
            'interface': dumps_json(interfaces) if interfaces else None,
            'license_start': _to_date(_c('建设时间')),
            'license_expiry': _to_date(_c('授权到期时间')),
            'os_version': _c('系统版本'),
            'rule_version': _c('规则库版本'),
            'is_maintenance': _c('硬件维修情况') not in ('', '无维修', '无'),
            'is_in_use': _c('是否在用') == '在用',
        })
    return rows, groups_skipped, errors


def resolve_customer(customer_id=None, create_if_missing=False):
    """按名称解析客户；缺失时可选创建（信息来自客户档案）"""
    if customer_id:
        c = Customer.query.get(customer_id)
        if not c:
            raise RuntimeError(f'客户 ID {customer_id} 不存在')
        return c
    c = Customer.query.filter_by(name=CUSTOMER_NAME).first()
    if c:
        return c
    if not create_if_missing:
        raise RuntimeError(f'客户「{CUSTOMER_NAME}」不存在（已按资产表名称匹配）。'
                           f'加 --create-customer 自动创建，或 --customer-id <id> 指定')
    # 匹配地区：南昌市 → 红谷滩区
    region = None
    for r in Region.query.filter(Region.name.contains('红谷滩')).all():
        if r.parent_id:
            parent = Region.query.get(r.parent_id)
            if parent and '南昌' in (parent.name or ''):
                region = r
                break
    from services.customer_service import create_customer
    c = create_customer({
        'name': CUSTOMER_NAME,
        'contact_person': '王科长',
        'phone': '13361718333',
        'address': '南昌市红谷滩区绿茵路1号南昌广电大楼9层',
        'region_id': region.id if region else None,
        'level': '常规',
    })
    db.session.commit()
    print(f'[客户] 已创建 {CUSTOMER_NAME} (id={c.id}, region={region.name if region else "未匹配"})')
    return c


def import_devices(path, customer_id=None, update=False, create_customer=False, dry_run=True):
    rows, groups_skipped, errors = parse_asset_excel(path)
    try:
        c = resolve_customer(customer_id, create_customer and not dry_run)
    except RuntimeError as e:
        if dry_run:
            print(f'[警告] {e}（预览模式继续展示解析结果；实际导入时请 --create-customer 或指定 --customer-id）')
            c = None
        else:
            raise
    if c is None:
        for r in rows:
            print(f"  [preview] {r['device_type'] or '其他':<6} {r['device_name']:<28} "
                  f"IP={r['ip_address'] or '-':<16} {r['location']}")
        if errors:
            print('\n[跳过/错误]')
            for e in errors:
                print(f'  ! {e}')
        print(f"\n共解析 {len(rows)} 台设备（客户未解析，未落库）。")
        return len(rows), 0, 0
    # db_existing 为导入前库中快照（只读，用于 skip/update 判定）；本批新建只进 seen_names
    db_existing = {d.device_name: d for d in Device.query.filter_by(customer_id=c.id).all()}
    created = updated = skipped = 0
    seen_names = set(db_existing)
    for r in rows:
        name = r['device_name']
        dev = db_existing.get(name)
        if dev:
            if not update:
                print(f"  [skip] 「{name}」（已存在 id={dev.id}；--update 可覆盖更新）")
                skipped += 1
                seen_names.add(name)
                continue
            # --update：直接更新库中原名设备，不改名
            if not dry_run:
                for k, v in r.items():
                    setattr(dev, k, v)
            seen_names.add(name)
            updated += 1
            print(f"  [update] {r['device_type'] or '其他':<6} {name:<28} "
                  f"IP={r['ip_address'] or '-':<16} {r['location']}")
            continue
        # 本批内重名（如多台「服务器」「交换机(24口)」跨机柜）自动加机柜后缀避免互相覆盖；
        # 库中已存在且 skip 的原始名占位，后续同名行才加后缀
        if name in seen_names:
            suffix = f'（{group_of(r["location"])}）'
            candidate = f'{name}{suffix}'
            n = 2
            while candidate in seen_names:
                candidate = f'{name}{n}{suffix}'
                n += 1
            name = candidate
            r = dict(r, device_name=name)
            dev = db_existing.get(name)
            if dev and not update:
                print(f"  [skip] 「{name}」（已存在 id={dev.id}）")
                skipped += 1
                seen_names.add(name)
                continue
            if dev:  # --update 且后缀名已存在：更新该台
                if not dry_run:
                    for k, v in r.items():
                        setattr(dev, k, v)
                seen_names.add(name)
                updated += 1
                print(f"  [update] {r['device_type'] or '其他':<6} {name:<28} "
                      f"IP={r['ip_address'] or '-':<16} {r['location']}")
                continue
        if not dry_run:
            db.session.add(Device(customer_id=c.id, **r))
        seen_names.add(name)
        created += 1
        print(f"  [create] {r['device_type'] or '其他':<6} {name:<28} "
              f"IP={r['ip_address'] or '-':<16} {r['location']}")
    if errors:
        print('\n[跳过/错误]')
        for e in errors:
            print(f'  ! {e}')
    if not dry_run and (created or updated):
        db.session.commit()
        from services.device_service import sync_customer_device_count
        sync_customer_device_count(c.id)
    print(f"\n客户: {c.name} (id={c.id}) | 行数: {len(rows)} | 创建: {created} | "
          f"更新: {updated} | 跳过: {skipped} | 组头: {groups_skipped} | 错误: {len(errors)}")
    if dry_run:
        print('以上为预览（未写库）。确认后执行: python scripts/import_asset_devices.py <xlsx> --apply')
    return created, updated, skipped


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    apply_flag = '--apply' in sys.argv
    update_flag = '--update' in sys.argv
    create_flag = '--create-customer' in sys.argv
    customer_id = None
    if '--customer-id' in sys.argv:
        i = sys.argv.index('--customer-id')
        customer_id = int(sys.argv[i + 1])
    if not args:
        print(__doc__)
        sys.exit(1)
    app = create_app()
    with app.app_context():
        try:
            import_devices(args[0], customer_id=customer_id,
                           update=update_flag, create_customer=create_flag,
                           dry_run=not apply_flag)
        except (ValueError, RuntimeError) as e:
            print(f'错误: {e}')
            sys.exit(1)


if __name__ == '__main__':
    main()
