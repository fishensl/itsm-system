# -*- coding: utf-8 -*-
"""资产管理蓝图：设备 CRUD / 详情 / 导入 / 导出 / API

业务规则下沉到 services/device_service.py。
拓扑/配置备份/采集（device_collect, config_backups, topologies）由于
业务复杂、模板互相引用多，暂留 app.py 中。
"""
import os
import tempfile
from datetime import date
from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, send_from_directory, jsonify, current_app)
from flask_login import login_required, current_user
from models import (Device, Customer, PasswordHistory, db, DeviceType, Brand,
                    NetworkType, CustomField, Region, DeviceFirmware,
                    DeviceConfigBackup)
from sqlalchemy.orm import joinedload
from utils.pagination import paginate, paginate_render_args
from utils.crypto import decrypt_password
from services.device_service import (create_device_from_form, update_device_from_form,
                                      delete_device)
from utils.upload import validate_upload, save_temp_upload, open_excel, cleanup_temp_file, ALLOWED_EXCEL_EXT, MAX_IMPORT_ROWS
from utils.permission import require_permission
from utils.decorators import api_view


asset_bp = Blueprint('asset', __name__)


# ============================ 设备列表 ============================
@asset_bp.route('/devices')
@login_required
@require_permission('device:view')
def device_list():
    model_filter = request.args.get('model', '')
    brand_filter = request.args.get('brand', '')
    type_filter = request.args.get('device_type', '')
    customer_filter = request.args.get('customer_id', '', type=int)
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    query = Device.query
    if search:
        query = query.filter(
            Device.device_name.contains(search) |
            Device.ip_address.contains(search) |
            Device.brand.contains(search)
        )
    if model_filter:
        query = query.filter(Device.model == model_filter)
    if brand_filter:
        query = query.filter(Device.brand == brand_filter)
    if type_filter:
        query = query.filter(Device.device_type == type_filter)
    if customer_filter:
        query = query.filter(Device.customer_id == customer_filter)
    # 预加载关联：customer / region_rel（详情/地区列） / region.parent（地区名拼接）
    query = query.options(
        joinedload(Device.customer).joinedload(Customer.region_rel).joinedload(Region.parent),
        joinedload(Device.region_rel).joinedload(Region.parent),
    )
    query = query.order_by(Device.id.desc())
    pag = paginate(query, page=page)
    customers = Customer.query.order_by(Customer.name).all()
    models_list = db.session.query(Device.model).distinct().filter(Device.model != '').all()
    brands_list = db.session.query(Device.brand).distinct().filter(Device.brand != '').all()
    types_list = db.session.query(Device.device_type).distinct().filter(Device.device_type != '').all()
    # 当前页按客户分组，再按父子单位嵌套（不再按地市分组 —— 客户名里已含市/县）
    # 父子关系来自 services/customer_hierarchy.derive_parent_id：
    #   Customer.parent_id（手动覆盖）> 同类别+父市的市级客户
    from services.customer_hierarchy import derive_parent_id, build_parent_index
    # 1. 当前页设备按客户分组
    by_customer = {}
    for d in pag['items']:
        if d.license_expiry:
            d._days_to_expiry = (d.license_expiry - date.today()).days
        else:
            d._days_to_expiry = None
        by_customer.setdefault(d.customer_id, []).append(d)
    # 2. 取出本页所涉客户（用于决定父子归属）。父客户即便本页无设备，也要拉进来。
    cust_ids = [cid for cid in by_customer.keys() if cid is not None]
    cust_map = {}
    if cust_ids:
        page_customers = Customer.query.options(
            joinedload(Customer.region_rel).joinedload(Region.parent),
            joinedload(Customer.category_rel),
        ).filter(Customer.id.in_(cust_ids)).all()
        extra_parent_keys = set()
        for c in page_customers:
            if c.region_rel and c.region_rel.parent_id and c.category_id:
                extra_parent_keys.add((c.region_rel.parent_id, c.category_id))
        extra_parents = []
        if extra_parent_keys:
            from sqlalchemy import and_, or_
            cond = None
            for rid, catid in extra_parent_keys:
                term = and_(Customer.region_id == rid, Customer.category_id == catid,
                            ~Customer.id.in_(cust_ids))
                cond = term if cond is None else or_(cond, term)
            if cond is not None:
                extra_parents = Customer.query.options(
                    joinedload(Customer.region_rel).joinedload(Region.parent),
                    joinedload(Customer.category_rel),
                ).filter(cond).all()
        for c in list(page_customers) + list(extra_parents):
            cust_map[c.id] = c
    # 3. 父子映射
    parent_index = build_parent_index(list(cust_map.values()))
    parent_of = {}
    children_of = {}
    for c in cust_map.values():
        pid = derive_parent_id(c, parent_index)
        if pid and pid in cust_map:
            parent_of[c.id] = pid
            children_of.setdefault(pid, []).append(c.id)
    # 4. 平铺为顶层节点列表（父行 → children），按客户名排序
    customer_groups = []
    top_ids = sorted([cid for cid in cust_map.keys() if cid not in parent_of],
                     key=lambda i: cust_map[i].name)
    for top_id in top_ids:
        c = cust_map[top_id]
        node = {
            'customer': c,
            'devices': by_customer.get(top_id, []),
            'children': [
                {'customer': cust_map[child_id], 'devices': by_customer.get(child_id, []),
                 'children': []}
                for child_id in sorted(children_of.get(top_id, []),
                                       key=lambda i: cust_map[i].name)
            ],
        }
        # 父客户若本页既无设备又无子单位的设备 → 跳过
        if not node['devices'] and not any(ch['devices'] for ch in node['children']):
            continue
        customer_groups.append(node)
    # 「未分配客户」兜底：customer_id 为 None 的设备
    if None in by_customer:
        customer_groups.append({
            'customer': None,
            'devices': by_customer[None],
            'children': [],
        })
    return render_template(
        'devices/list.html', **paginate_render_args(pag),
        customers=customers,
        all_customers=customers,
        regions=Region.query.order_by(Region.parent_id.is_(None).desc(), Region.sort_order, Region.id).all(),
        device_types=DeviceType.query.order_by(DeviceType.sort_order, DeviceType.id).all(),
        brands=Brand.query.order_by(Brand.sort_order, Brand.id).all(),
        models_list=[m[0] for m in models_list if m[0]],
        brands_list=[b[0] for b in brands_list if b[0]],
        types_list=[t[0] for t in types_list if t[0]],
        customer_groups=customer_groups,
        filters={
            'model': model_filter, 'brand': brand_filter,
            'device_type': type_filter, 'customer_id': customer_filter,
            'search': search,
        }
    )


# ============================ 设备新增 ============================
@asset_bp.route('/devices/add', methods=['POST'])
@login_required
@require_permission('device:add')
def device_add():
    try:
        d = create_device_from_form(request.form)
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('设备添加失败')
        flash(str(e) or '设备添加失败', 'danger')
        return redirect(url_for('asset.device_list'))
    _sync_customer_device_count(d.customer_id)
    flash('设备添加成功', 'success')
    return redirect(url_for('asset.device_list'))


@asset_bp.route('/devices/edit-page/<int:id>', methods=['GET'])
@login_required
@require_permission('device:edit')
def device_edit_page(id):
    # 编辑已改为 AJAX 弹窗，此页面仅保留兼容重定向
    return redirect(url_for('asset.device_list'))


@asset_bp.route('/devices/edit/<int:id>', methods=['POST'])
@login_required
@require_permission('device:edit')
def device_edit(id):
    try:
        form = request.form.copy()
        form['changed_by_name'] = current_user.realname or current_user.username
        d = update_device_from_form(id, form)
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('设备编辑失败')
        flash(str(e) or '设备更新失败', 'danger')
        return redirect(url_for('asset.device_list'))
    new_cid = d.customer_id
    _sync_customer_device_count(new_cid)
    flash('设备信息已更新', 'success')
    return redirect(url_for('asset.device_list'))


@asset_bp.route('/devices/delete/<int:id>', methods=['POST'])
@login_required
@require_permission('device:delete')
def device_delete(id):
    try:
        cid = delete_device(id)
    except Exception as e:
        db.session.rollback()
        flash(str(e) or '设备删除失败', 'danger')
        return redirect(url_for('asset.device_list'))
    _sync_customer_device_count(cid)
    flash('设备已删除', 'success')
    return redirect(url_for('asset.device_list'))


# ============================ 设备详情 ============================
@asset_bp.route('/devices/<int:id>')
@login_required
@require_permission('device:view')
def device_detail(id):
    import json as _json
    d = Device.query.get_or_404(id)
    customer = Customer.query.get(d.customer_id) if d.customer_id else None
    password = decrypt_password(d.password_encrypted) if d.password_encrypted else ''
    remaining = (d.license_expiry - date.today()).days if d.license_expiry else None
    interface_list = _json.loads(d.interface) if d.interface and d.interface.startswith('[') else (
        [d.interface] if d.interface else []
    )
    histories = PasswordHistory.query.filter_by(device_id=id)\
        .order_by(PasswordHistory.id.desc()).limit(20).all()
    return render_template('devices/detail.html',
                           device=d, customer=customer, password=password,
                           remaining=remaining, interface_list=interface_list,
                           histories=histories)


# ============================ API: 设备 JSON ============================
@asset_bp.route('/api/devices/<int:id>')
@login_required
@require_permission('device:view')
@api_view
def api_device_get(id):
    import json as _json
    d = Device.query.get_or_404(id)
    return jsonify({
        'id': d.id,
        'customer_id': d.customer_id,
        'region_id': d.region_id,
        'device_name': d.device_name,
        'device_type': d.device_type,
        'brand': d.brand,
        'model': d.model,
        'serial_number': d.serial_number or '',
        'ip_address': d.ip_address,
        'port': d.port,
        'username': d.username,
        'password': decrypt_password(d.password_encrypted) if d.password_encrypted else '',
        'login_method': d.login_method,
        'location': d.location,
        'interface': _json.loads(d.interface) if d.interface and d.interface.startswith('[') else (
            [d.interface] if d.interface else []
        ),
        'os_version': d.os_version,
        'rule_version': d.rule_version,
        'is_maintenance': d.is_maintenance,
        'is_in_use': d.is_in_use,
        'license_expiry': d.license_expiry.strftime('%Y-%m-%d') if d.license_expiry else '',
        'license_start': d.license_start.strftime('%Y-%m-%d') if d.license_start else '',
        'remark': d.remark,
    })


# ============================ 设备导入/导出 ============================
@asset_bp.route('/devices/export', methods=['POST'])
@login_required
@require_permission('device:view')
def device_export():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    search = request.args.get('search', '')
    customer_filter = request.args.get('customer_id', '', type=int)
    query = Device.query
    if search:
        query = query.filter(
            Device.device_name.contains(search) |
            Device.ip_address.contains(search) |
            Device.brand.contains(search)
        )
    if customer_filter:
        query = query.filter(Device.customer_id == customer_filter)
    devices = query.order_by(Device.id.desc()).all()
    selected_cols = request.form.getlist('export_columns')
    all_columns = {
        'customer_name': '所属客户', 'device_name': '设备名称', 'device_type': '设备类型',
        'brand': '品牌', 'model': '型号', 'serial_number': '序列号', 'ip_address': 'IP地址',
        'port': '端口', 'username': '登录用户名', 'password': '登录密码',
        'license_expiry': '授权截止日期', 'license_start': '授权开始日期', 'login_method': '登录方式', 'location': '安装位置',
        'os_version': '系统版本', 'rule_version': '规则库版本',
        'is_maintenance': '是否维修', 'is_in_use': '是否在用',
        'license_remaining_days': '剩余天数', 'remark': '备注',
    }
    if not selected_cols:
        selected_cols = list(all_columns.keys())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '设备信息'
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1890FF', end_color='096DD9', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    headers = [all_columns[c] for c in selected_cols]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = thin_border
    for row_idx, d in enumerate(devices, 2):
        license_remaining = (d.license_expiry - date.today()).days if d.license_expiry else ''
        data_map = {
            'customer_name': d.customer.name if d.customer else '',
            'device_name': d.device_name, 'device_type': d.device_type,
            'brand': d.brand, 'model': d.model,
            'serial_number': d.serial_number or '',
            'ip_address': d.ip_address, 'port': d.port,
            'username': d.username,
            'password': decrypt_password(d.password_encrypted) if d.password_encrypted else '',
            'license_expiry': d.license_expiry.strftime('%Y-%m-%d') if d.license_expiry else '',
            'license_start': d.license_start.strftime('%Y-%m-%d') if d.license_start else '',
            'login_method': d.login_method, 'location': d.location or '',
            'os_version': d.os_version or '', 'rule_version': d.rule_version or '',
            'is_maintenance': '是' if d.is_maintenance else '否',
            'is_in_use': '是' if d.is_in_use else '否',
            'license_remaining_days': license_remaining,
            'remark': d.remark or '',
        }
        for col_idx, c in enumerate(selected_cols, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=data_map.get(c, ''))
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()
    return send_from_directory(
        os.path.dirname(tmp.name), os.path.basename(tmp.name),
        as_attachment=True,
        download_name=f'设备导出_{date.today().isoformat()}.xlsx'
    )


@asset_bp.route('/devices/import', methods=['POST'])
@login_required
@require_permission('device:add')
def device_import():
    """批量导入设备信息"""
    if 'import_file' not in request.files:
        flash('请选择要导入的 Excel 文件', 'danger')
        return redirect(url_for('asset.device_list'))
    f = request.files['import_file']
    ok, err, _ = validate_upload(f, ALLOWED_EXCEL_EXT, max_size_mb=20)
    if not ok:
        flash(err, 'danger')
        return redirect(url_for('asset.device_list'))
    tmp = save_temp_upload(f, suffix='.xlsx')
    success_count = 0
    error_count = 0
    errors = []
    try:
        wb, ws, err = open_excel(tmp, app=current_app)
        if err:
            flash(err[0], err[1])
            return redirect(url_for('asset.device_list'))

        header_row = [cell.value for cell in ws[1]]
        col_map = {}
        for idx, h in enumerate(header_row):
            if h:
                col_map[str(h).strip()] = idx

        field_mapping = {
            '所属客户': 'customer_name', '设备名称': 'device_name', '设备类型': 'device_type',
            '品牌': 'brand', '型号': 'model', '序列号': 'serial_number', 'IP地址': 'ip_address',
            '端口': 'port', '登录用户名': 'username', '登录密码': 'password',
            '授权截止日期': 'license_expiry', '授权开始日期': 'license_start', '登录方式': 'login_method', '安装位置': 'location',
            '系统版本': 'os_version', '规则库版本': 'rule_version', '备注': 'remark',
        }

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for cn, idx in col_map.items():
                val = ws.cell(row=row_idx, column=idx + 1).value
                field = field_mapping.get(cn)
                if field:
                    row_data[field] = str(val).strip() if val else ''

            device_name = row_data.get('device_name', '')
            if not device_name:
                error_count += 1
                errors.append(f'第{row_idx}行：设备名称为空，跳过')
                continue

            customer = None
            if 'customer_name' in row_data and row_data['customer_name']:
                customer = Customer.query.filter_by(name=row_data['customer_name']).first()
                if not customer:
                    flash(f'客户 "{row_data["customer_name"]}" 不存在', 'warning')
            try:
                from services.device_service import _parse_date
                from utils.crypto import encrypt_password as _ep
                plain_password = row_data.get('password', '')
                encrypted = _ep(plain_password) if plain_password else ''
                license_expiry = _parse_date(row_data.get('license_expiry'))
                license_start = _parse_date(row_data.get('license_start'))

                d = Device(
                    customer_id=customer.id if customer else None,
                    device_name=device_name,
                    device_type=row_data.get('device_type', ''),
                    brand=row_data.get('brand', ''),
                    model=row_data.get('model', ''),
                    serial_number=row_data.get('serial_number', ''),
                    ip_address=row_data.get('ip_address', ''),
                    port=int(row_data.get('port', 22)) if row_data.get('port') else 22,
                    username=row_data.get('username', ''),
                    password_encrypted=encrypted,
                    login_method=row_data.get('login_method', ''),
                    os_version=row_data.get('os_version', ''),
                    rule_version=row_data.get('rule_version', ''),
                    is_maintenance=row_data.get('is_maintenance', '') in ('是', '1', 'true', 'True'),
                    is_in_use=row_data.get('is_in_use', '') in ('是', '1', 'true', 'True'),
                    license_expiry=license_expiry,
                    license_start=license_start,
                    remark=row_data.get('remark', ''),
                )
                db.session.add(d)
                db.session.commit()
                success_count += 1
            except Exception as e:
                db.session.rollback()
                error_count += 1
                errors.append(f'第{row_idx}行（{device_name}）：{e}')

        msg = f'导入完成：成功 {success_count} 条'
        if error_count:
            msg += f'，失败 {error_count} 条'
            for err in errors[:5]:
                flash(err, 'danger')
        flash(msg, 'success' if success_count else 'danger')
    finally:
        cleanup_temp_file(tmp)
    return redirect(url_for('asset.device_list'))


# ============================ 内部工具 ============================
def _sync_customer_device_count(customer_id):
    """同步客户的 device_count 冗余字段（蓝图内部使用）"""
    if not customer_id:
        return
    from models import Customer
    from app import calculate_customer_tier
    cnt = Device.query.filter_by(customer_id=customer_id).count()
    c = Customer.query.get(customer_id)
    if c:
        c.device_count = cnt
        auto_tier = calculate_customer_tier(cnt, c.has_onsite, c.has_drill)
        if c.level not in ('核心', '重点', '常规') or not c.level:
            c.level = auto_tier
        db.session.commit()


# ============================ 设备配置子路由 ============================
@asset_bp.route("/device-types", methods=["GET", "POST"])
@login_required
@require_permission("device:view")
def device_type_list():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if name:
            dt = DeviceType(name=name, sort_order=int(request.form.get("sort_order") or 0))
            db.session.add(dt); db.session.commit()
            flash("已添加", "success")
        return redirect(url_for("asset.device_type_list"))
    types = DeviceType.query.order_by(DeviceType.sort_order, DeviceType.id).all()
    return render_template("device_types/list.html", types=types)


@asset_bp.route("/device-types/edit/<int:id>", methods=["POST"])
@login_required
@require_permission("device:edit")
def device_type_edit(id):
    dt = DeviceType.query.get_or_404(id)
    dt.name = (request.form.get("name") or dt.name).strip()
    dt.sort_order = int(request.form.get("sort_order") or 0)
    db.session.commit()
    flash("已更新", "success")
    return redirect(url_for("asset.device_type_list"))


@asset_bp.route("/device-types/delete/<int:id>", methods=['POST'])
@login_required
@require_permission("device:delete")
def device_type_delete(id):
    DeviceType.query.filter_by(id=id).delete()
    db.session.commit()
    flash("已删除", "success")
    return redirect(url_for("asset.device_type_list"))


@asset_bp.route("/device-brands", methods=["GET", "POST"])
@login_required
@require_permission("device:view")
def brand_list():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        sort_order = int(request.form.get("sort_order") or 0)
        if name:
            b = Brand(name=name, sort_order=sort_order)
            db.session.add(b); db.session.commit()
            flash("已添加", "success")
        return redirect(url_for("asset.brand_list"))
    brands = Brand.query.order_by(Brand.sort_order, Brand.id).all()
    return render_template("brands/list.html", brands=brands)


@asset_bp.route("/device-brands/edit/<int:id>", methods=["POST"])
@login_required
@require_permission("device:edit")
def brand_edit(id):
    b = Brand.query.get_or_404(id)
    b.name = (request.form.get("name") or b.name).strip()
    if request.form.get("sort_order") is not None:
        b.sort_order = int(request.form.get("sort_order") or 0)
    db.session.commit()
    flash("已更新", "success")
    return redirect(url_for("asset.brand_list"))


@asset_bp.route("/device-brands/delete/<int:id>", methods=['POST'])
@login_required
@require_permission("device:delete")
def brand_delete(id):
    Brand.query.filter_by(id=id).delete()
    db.session.commit()
    flash("已删除", "success")
    return redirect(url_for("asset.brand_list"))


@asset_bp.route("/device-network-types", methods=["GET", "POST"])
@login_required
@require_permission("device:view")
def network_type_list():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if name:
            n = NetworkType(name=name)
            db.session.add(n); db.session.commit()
            flash("已添加", "success")
        return redirect(url_for("asset.network_type_list"))
    types = NetworkType.query.order_by(NetworkType.id).all()
    return render_template("network_types/list.html", types=types)


@asset_bp.route("/device-network-types/edit/<int:id>", methods=["POST"])
@login_required
@require_permission("device:edit")
def network_type_edit(id):
    n = NetworkType.query.get_or_404(id)
    n.name = (request.form.get("name") or n.name).strip()
    db.session.commit()
    flash("已更新", "success")
    return redirect(url_for("asset.network_type_list"))


@asset_bp.route("/device-network-types/delete/<int:id>", methods=['POST'])
@login_required
@require_permission("device:delete")
def network_type_delete(id):
    NetworkType.query.filter_by(id=id).delete()
    db.session.commit()
    flash("已删除", "success")
    return redirect(url_for("asset.network_type_list"))


@asset_bp.route("/device-custom-fields", methods=["GET", "POST"])
@login_required
@require_permission("device:view")
def custom_field_list():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if name:
            f = CustomField(name=name, field_type=request.form.get("field_type", "text"))
            db.session.add(f); db.session.commit()
            flash("已添加", "success")
        return redirect(url_for("asset.custom_field_list"))
    fields = CustomField.query.order_by(CustomField.id).all()
    return render_template("custom_fields/list.html", fields=fields)


@asset_bp.route("/device-custom-fields/edit/<int:id>", methods=["POST"])
@login_required
@require_permission("device:edit")
def custom_field_edit(id):
    f = CustomField.query.get_or_404(id)
    f.name = (request.form.get("name") or f.name).strip()
    f.field_type = request.form.get("field_type", f.field_type)
    db.session.commit()
    flash("已更新", "success")
    return redirect(url_for("asset.custom_field_list"))


@asset_bp.route("/device-custom-fields/delete/<int:id>", methods=['POST'])
@login_required
@require_permission("device:delete")
def custom_field_delete(id):
    CustomField.query.filter_by(id=id).delete()
    db.session.commit()
    flash("已删除", "success")
    return redirect(url_for("asset.custom_field_list"))


# ============================ 设备类型 add（之前漏了） ============================
@asset_bp.route("/device-types/add", methods=["POST"])
@login_required
@require_permission("device:edit")
def device_type_add():
    name = (request.form.get("name") or "").strip()
    if name:
        dt = DeviceType(name=name, sort_order=int(request.form.get("sort_order") or 0))
        db.session.add(dt); db.session.commit()
        flash("已添加", "success")
    return redirect(url_for("asset.device_type_list"))


# ============================ 品牌 add（之前漏了） ============================
@asset_bp.route("/device-brands/add", methods=["POST"])
@login_required
@require_permission("device:edit")
def brand_add():
    name = (request.form.get("name") or "").strip()
    sort_order = int(request.form.get("sort_order") or 0)
    if name:
        b = Brand(name=name, sort_order=sort_order)
        db.session.add(b); db.session.commit()
        flash("已添加", "success")
    return redirect(url_for("asset.brand_list"))


# ============================ 网络类型 add ============================
@asset_bp.route("/device-network-types/add", methods=["POST"])
@login_required
@require_permission("device:edit")
def network_type_add():
    name = (request.form.get("name") or "").strip()
    if name:
        n = NetworkType(name=name)
        db.session.add(n); db.session.commit()
        flash("已添加", "success")
    return redirect(url_for("asset.network_type_list"))


# ============================ 自定义字段 add ============================
@asset_bp.route("/device-custom-fields/add", methods=["POST"])
@login_required
@require_permission("device:edit")
def custom_field_add():
    name = (request.form.get("name") or "").strip()
    if name:
        f = CustomField(name=name, field_type=request.form.get("field_type", "text"))
        db.session.add(f); db.session.commit()
        flash("已添加", "success")
    return redirect(url_for("asset.custom_field_list"))


# ============================ 设备固件版本库 (V12) ============================
def _parse_date_arg(s):
    if not s: return None
    try:
        return date.fromisoformat(s.strip())
    except Exception:
        return None


@asset_bp.route('/device-firmwares')
@login_required
@require_permission('device:view')
def firmware_list():
    """固件版本库列表 — 按 brand+model 分组展示，附带使用该型号的设备列表（用于版本对比）"""
    brand_filter = request.args.get('brand', '')
    model_filter = request.args.get('model', '')
    type_filter = request.args.get('firmware_type', '')

    q = DeviceFirmware.query
    if brand_filter:
        q = q.filter(DeviceFirmware.brand == brand_filter)
    if model_filter:
        q = q.filter(DeviceFirmware.model == model_filter)
    if type_filter:
        q = q.filter(DeviceFirmware.firmware_type == type_filter)
    firmwares = q.order_by(
        DeviceFirmware.brand, DeviceFirmware.model,
        DeviceFirmware.firmware_type, DeviceFirmware.is_latest.desc(),
        DeviceFirmware.release_date.desc()
    ).all()

    # 按 (brand, model) 分组：每组下再按 firmware_type 分组
    from collections import OrderedDict
    grouped = OrderedDict()
    for fw in firmwares:
        key = (fw.brand or '未分类', fw.model or '未分类型号')
        grouped.setdefault(key, OrderedDict()).setdefault(fw.firmware_type or '其他', []).append(fw)

    # 为每组挂上设备清单（同 brand+model 的所有设备，便于对比版本）
    group_devices = {}
    for (brand, model) in grouped.keys():
        devs = Device.query.filter_by(brand=brand, model=model).all()
        group_devices[(brand, model)] = devs

    # 筛选下拉
    all_brands = sorted(set(b for b, _ in grouped.keys() if b))
    all_models = sorted(set(m for _, m in grouped.keys() if m))
    all_types = ['系统固件', '规则库', 'BIOS', '其他']

    return render_template('device_firmwares/list.html',
                           grouped=grouped, group_devices=group_devices,
                           all_brands=all_brands, all_models=all_models, all_types=all_types,
                           brand_filter=brand_filter, model_filter=model_filter, type_filter=type_filter)


@asset_bp.route('/device-firmwares/add', methods=['POST'])
@login_required
@require_permission('device:edit')
def firmware_add():
    is_latest = request.form.get('is_latest') == 'on'
    fw = DeviceFirmware(
        brand=(request.form.get('brand') or '').strip(),
        model=(request.form.get('model') or '').strip(),
        firmware_type=request.form.get('firmware_type', '系统固件'),
        version=(request.form.get('version') or '').strip(),
        release_date=_parse_date_arg(request.form.get('release_date')),
        changelog=request.form.get('changelog', ''),
        download_url=request.form.get('download_url', ''),
        file_size_mb=float(request.form.get('file_size_mb') or 0) if request.form.get('file_size_mb') else 0,
        md5_checksum=request.form.get('md5_checksum', ''),
        is_latest=is_latest,
        min_compatible_hardware=request.form.get('min_compatible_hardware', ''),
        upgrade_guide=request.form.get('upgrade_guide', ''),
        remark=request.form.get('remark', ''),
    )
    if not fw.brand or not fw.model or not fw.version:
        flash('品牌/型号/版本号为必填项', 'danger')
        return redirect(url_for('asset.firmware_list'))
    # 同 brand+model+firmware_type 仅一条 is_latest=True
    if is_latest:
        DeviceFirmware.query.filter_by(brand=fw.brand, model=fw.model, firmware_type=fw.firmware_type
                                       ).update({'is_latest': False})
    db.session.add(fw); db.session.commit()
    flash('已添加固件版本', 'success')
    return redirect(url_for('asset.firmware_list'))


@asset_bp.route('/device-firmwares/edit/<int:id>', methods=['POST'])
@login_required
@require_permission('device:edit')
def firmware_edit(id):
    fw = DeviceFirmware.query.get_or_404(id)
    fw.brand = (request.form.get('brand') or fw.brand).strip()
    fw.model = (request.form.get('model') or fw.model).strip()
    fw.firmware_type = request.form.get('firmware_type', fw.firmware_type)
    fw.version = (request.form.get('version') or fw.version).strip()
    fw.release_date = _parse_date_arg(request.form.get('release_date')) or fw.release_date
    fw.changelog = request.form.get('changelog', fw.changelog)
    fw.download_url = request.form.get('download_url', fw.download_url)
    try:
        if request.form.get('file_size_mb'):
            fw.file_size_mb = float(request.form['file_size_mb'])
    except (TypeError, ValueError):
        pass
    fw.md5_checksum = request.form.get('md5_checksum', fw.md5_checksum)
    fw.min_compatible_hardware = request.form.get('min_compatible_hardware', fw.min_compatible_hardware)
    fw.upgrade_guide = request.form.get('upgrade_guide', fw.upgrade_guide)
    fw.remark = request.form.get('remark', fw.remark)
    new_is_latest = request.form.get('is_latest') == 'on'
    if new_is_latest and not fw.is_latest:
        # 切换为最新 → 清掉同组其他 latest
        DeviceFirmware.query.filter(
            DeviceFirmware.brand == fw.brand,
            DeviceFirmware.model == fw.model,
            DeviceFirmware.firmware_type == fw.firmware_type,
            DeviceFirmware.id != fw.id,
        ).update({'is_latest': False})
    fw.is_latest = new_is_latest
    db.session.commit()
    flash('已更新', 'success')
    return redirect(url_for('asset.firmware_list'))


@asset_bp.route('/device-firmwares/delete/<int:id>', methods=['POST'])
@login_required
@require_permission('device:delete')
def firmware_delete(id):
    fw = DeviceFirmware.query.get(id)
    if fw:
        db.session.delete(fw); db.session.commit()
    flash('已删除', 'success')
    return redirect(url_for('asset.firmware_list'))


@asset_bp.route('/api/firmwares/match-device/<int:device_id>')
@login_required
@require_permission('device:view')
def api_firmware_match_device(device_id):
    """V12: 给定设备 id，返回该设备 brand+model 下所有固件版本（以及最新版本标记）"""
    d = Device.query.get_or_404(device_id)
    fws = DeviceFirmware.query.filter_by(brand=d.brand, model=d.model).order_by(
        DeviceFirmware.firmware_type, DeviceFirmware.is_latest.desc(), DeviceFirmware.release_date.desc()
    ).all()
    return jsonify({
        'device': {
            'id': d.id, 'name': d.device_name, 'brand': d.brand, 'model': d.model,
            'os_version': d.os_version or '', 'rule_version': d.rule_version or '',
        },
        'firmwares': [{
            'id': fw.id, 'firmware_type': fw.firmware_type, 'version': fw.version,
            'release_date': fw.release_date.isoformat() if fw.release_date else '',
            'is_latest': fw.is_latest,
            'changelog': fw.changelog, 'download_url': fw.download_url,
            'upgrade_guide': fw.upgrade_guide,
        } for fw in fws],
    })


# ============================ 配置备份（V14） ============================
def _compute_config_diff(text_a, text_b):
    """逐行差异：使用 difflib.SequenceMatcher 计算 equal/delete/insert/replace 标记，
    返回 [{tag, line_a, line_b}, ...]，与 templates/devices/config_backups.html 的渲染契约一致。"""
    import difflib
    lines_a = (text_a or '').splitlines()
    lines_b = (text_b or '').splitlines()
    result = []
    matcher = difflib.SequenceMatcher(None, lines_a, lines_b)
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'equal':
            for k in range(i2 - i1):
                result.append({'tag': 'equal', 'line_a': lines_a[i1 + k], 'line_b': lines_b[j1 + k]})
        elif tag == 'delete':
            for k in range(i1, i2):
                result.append({'tag': 'delete', 'line_a': lines_a[k], 'line_b': ''})
        elif tag == 'insert':
            for k in range(j1, j2):
                result.append({'tag': 'insert', 'line_a': '', 'line_b': lines_b[k]})
        elif tag == 'replace':
            max_len = max(i2 - i1, j2 - j1)
            for k in range(max_len):
                la = lines_a[i1 + k] if k < (i2 - i1) else ''
                lb = lines_b[j1 + k] if k < (j2 - j1) else ''
                result.append({'tag': 'replace', 'line_a': la, 'line_b': lb})
    return result


@asset_bp.route('/devices/<int:id>/config-backups')
@login_required
@require_permission('device:view')
def device_config_backups(id):
    d = Device.query.get_or_404(id)
    backups = DeviceConfigBackup.query.filter_by(device_id=id)\
        .order_by(DeviceConfigBackup.id.desc()).all()
    diff_lines = None
    a_id = request.args.get('a', type=int)
    b_id = request.args.get('b', type=int)
    if a_id and b_id and a_id != b_id:
        ba = DeviceConfigBackup.query.filter_by(id=a_id, device_id=id).first()
        bb = DeviceConfigBackup.query.filter_by(id=b_id, device_id=id).first()
        if ba and bb:
            diff_lines = _compute_config_diff(ba.config_content or '', bb.config_content or '')
    return render_template('devices/config_backups.html',
                           device=d, backups=backups,
                           diff_lines=diff_lines, a_id=a_id, b_id=b_id)


@asset_bp.route('/devices/<int:id>/config-backups/add', methods=['POST'])
@login_required
@require_permission('device:edit')
def device_config_backup_add(id):
    import hashlib
    from werkzeug.utils import secure_filename
    d = Device.query.get_or_404(id)
    content = request.form.get('config_content', '')
    backup_type = request.form.get('backup_type', '运行配置')
    backup_method = request.form.get('backup_method', '手动输入')

    # 文件上传（可选）
    file_path = ''
    f = request.files.get('config_file')
    if f and f.filename:
        upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'configs', str(id))
        os.makedirs(upload_dir, exist_ok=True)
        safe_name = secure_filename(f.filename) or 'config.txt'
        # 防止重名覆盖：附加时间戳
        from datetime import datetime as _dt
        ts = _dt.now().strftime('%Y%m%d_%H%M%S')
        name_base, name_ext = os.path.splitext(safe_name)
        safe_name = f'{name_base}_{ts}{name_ext}'
        full_path = os.path.join(upload_dir, safe_name)
        f.save(full_path)
        file_path = f'uploads/configs/{id}/{safe_name}'
        backup_method = '文件上传'
        # 若用户没填内容，从文件读取
        if not content:
            try:
                with open(full_path, 'r', encoding='utf-8', errors='replace') as fh:
                    content = fh.read()
            except Exception:
                pass

    if not content and not file_path:
        flash('请填写配置内容或上传配置文件', 'warning')
        return redirect(url_for('asset.device_config_backups', id=id))

    checksum = hashlib.md5(content.encode('utf-8')).hexdigest() if content else ''

    backup = DeviceConfigBackup(
        device_id=id,
        backup_type=backup_type,
        config_content=content,
        backup_method=backup_method,
        backup_date=date.today(),
        file_path=file_path,
        checksum=checksum,
        created_by=getattr(current_user, 'realname', None) or current_user.username,
    )
    db.session.add(backup)
    db.session.commit()
    flash('配置备份已保存', 'success')
    return redirect(url_for('asset.device_config_backups', id=id))


@asset_bp.route('/devices/<int:id>/config-backups/<int:bid>/delete', methods=['POST'])
@login_required
@require_permission('device:delete')
def device_config_backup_delete(id, bid):
    backup = DeviceConfigBackup.query.filter_by(id=bid, device_id=id).first_or_404()
    # 删关联文件
    if backup.file_path:
        full = os.path.join(current_app.root_path, 'static', backup.file_path.replace('/', os.sep))
        if os.path.exists(full):
            try:
                os.remove(full)
            except Exception:
                pass
    db.session.delete(backup)
    db.session.commit()
    flash('配置备份已删除', 'success')
    return redirect(url_for('asset.device_config_backups', id=id))


@asset_bp.route('/devices/<int:id>/config-backups/compare')
@login_required
@require_permission('device:view')
def device_config_backup_compare(id):
    """compareForm 使用 GET 提交两个 v 复选框值；这里转成 ?a=&b= 重定向到列表页统一渲染。"""
    vs = request.args.getlist('v')
    a = vs[0] if len(vs) >= 1 else request.args.get('a')
    b = vs[1] if len(vs) >= 2 else request.args.get('b')
    if a and b:
        return redirect(url_for('asset.device_config_backups', id=id, a=a, b=b))
    flash('请勾选两个版本进行对比', 'warning')
    return redirect(url_for('asset.device_config_backups', id=id))


@asset_bp.route('/devices/<int:id>/collect')
@login_required
@require_permission('device:edit')
def device_collect(id):
    """从设备采集（SSH/Telnet/SNMP）— 占位实现，提示用户使用文件上传。"""
    d = Device.query.get_or_404(id)
    flash('远程采集功能尚未启用，请使用「文件上传」或「手动输入」方式新增备份。', 'info')
    return redirect(url_for('asset.device_config_backups', id=d.id))


@asset_bp.route('/api/devices/<int:id>/config-backups/upload-from-inspection', methods=['POST'])
@login_required
@require_permission('device:edit')
@api_view
def api_config_backup_upload(id):
    """巡检表单中 config_backup 字段类型上传配置文件时调用，自动创建一条 DeviceConfigBackup 记录。"""
    import hashlib
    from werkzeug.utils import secure_filename
    Device.query.get_or_404(id)
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': '未选择文件'}), 400
    upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'configs', str(id))
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = secure_filename(f.filename) or 'config.txt'
    from datetime import datetime as _dt
    ts = _dt.now().strftime('%Y%m%d_%H%M%S')
    name_base, name_ext = os.path.splitext(safe_name)
    safe_name = f'{name_base}_{ts}{name_ext}'
    full_path = os.path.join(upload_dir, safe_name)
    f.save(full_path)
    try:
        with open(full_path, 'r', encoding='utf-8', errors='replace') as fh:
            content = fh.read()
    except Exception:
        content = ''
    checksum = hashlib.md5(content.encode('utf-8')).hexdigest() if content else ''
    version = (request.form.get('version') or '').strip()
    backup = DeviceConfigBackup(
        device_id=id,
        backup_type='运行配置',
        config_content=content,
        backup_method='巡检上传',
        backup_date=date.today(),
        file_path=f'uploads/configs/{id}/{safe_name}',
        checksum=checksum,
        created_by=(getattr(current_user, 'realname', None) or current_user.username) + (f' / {version}' if version else ''),
    )
    db.session.add(backup)
    db.session.commit()
    return jsonify({
        'success': True,
        'backup_id': backup.id,
        'checksum': checksum,
        'file_path': backup.file_path,
        'filename': safe_name,
    })
