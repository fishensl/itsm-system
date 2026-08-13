# -*- coding: utf-8 -*-
"""V24 设备密码导出审核流：状态机、一次性令牌、加密包密码（pyzipper 解密验证）、通知与审计"""
import io
import pytest

from models import db, Customer, Device, User, Notification, AuditLog, ExportFile, DeviceExportRequest
from utils.crypto import encrypt_password


@pytest.fixture()
def seed(app):
    with app.app_context():
        c = Customer(name='密码客户')
        db.session.add(c)
        db.session.flush()
        d = Device(customer_id=c.id, device_name='密码设备', ip_address='10.1.1.1',
                   username='root', password_encrypted=encrypt_password('p@ssw0rd!'),
                   login_method='SSH')
        db.session.add(d)
        op = User.query.filter_by(username='op').first()
        op.customers = [c]
        db.session.commit()
        yield {'c': c.id, 'd': d.id}


class TestRequest:
    def test_require_reason(self, op_client, seed):
        r = op_client.post('/api/v2/devices/export-password-request',
                           json={'filters': {'columns': ['name', 'password']}, 'reason': ''})
        assert r.status_code == 400

    def test_require_password_column(self, op_client, seed):
        r = op_client.post('/api/v2/devices/export-password-request',
                           json={'filters': {'columns': ['name']}, 'reason': '等保审计'})
        assert r.status_code == 400

    def test_request_requires_reveal_permission(self, viewer_client, seed):
        r = viewer_client.post('/api/v2/devices/export-password-request', json={
            'filters': {'preset': 'password', 'customer_id': seed['c']},
            'reason': 'not allowed',
        })
        assert r.status_code == 403

    def test_request_rejects_out_of_scope_customer(self, app, op_client, seed):
        with app.app_context():
            customer = Customer(name='scope-hidden-request-customer')
            db.session.add(customer)
            db.session.commit()
            customer_id = customer.id
        r = op_client.post('/api/v2/devices/export-password-request', json={
            'filters': {'preset': 'password', 'customer_id': customer_id},
            'reason': 'not allowed',
        })
        assert r.status_code == 404

    def test_submit_creates_pending_and_notifies_admin(self, app, op_client, seed):
        r = op_client.post('/api/v2/devices/export-password-request', json={
            'filters': {'preset': 'password', 'customer_id': seed['c']},
            'reason': '等保审计需要导出密码台账'})
        assert r.get_json()['code'] == 0
        req_id = r.get_json()['data']['id']
        with app.app_context():
            req = db.session.get(DeviceExportRequest, req_id)
            assert req.status == 'pending'
            assert req.user_id == User.query.filter_by(username='op').first().id
            # 通知 admin
            admin_ids = [u.id for u in User.query.filter_by(username='admin').all()]
            msgs = Notification.query.filter(Notification.user_id.in_(admin_ids)).all()
            assert any('导出申请' in m.title for m in msgs)
            # 审计
            aud = AuditLog.query.filter_by(action='device:export_request').all()
            assert aud

    def test_mine_list(self, app, op_client, admin_client, seed):
        op_client.post('/api/v2/devices/export-password-request', json={
            'filters': {'preset': 'password'}, 'reason': '审计用途'})
        r = op_client.get('/api/v2/devices/export-password-requests?scope=mine')
        assert r.get_json()['code'] == 0
        items = r.get_json()['data']['items']
        assert len(items) == 1
        assert items[0]['status'] == 'pending'
        assert items[0]['realname'] == 'op'
        # admin 看自己的没有
        r = admin_client.get('/api/v2/devices/export-password-requests?scope=mine')
        assert r.get_json()['data']['items'] == []

    def test_review_list_admin_only(self, op_client, viewer_client, seed):
        r = op_client.get('/api/v2/devices/export-password-reviews')
        assert r.status_code == 403
        r = viewer_client.get('/api/v2/devices/export-password-reviews')
        assert r.status_code == 403


class TestReviewFlow:
    def _submit(self, op_client, seed):
        r = op_client.post('/api/v2/devices/export-password-request', json={
            'filters': {'preset': 'password', 'customer_id': seed['c']},
            'reason': '等保审计'})
        return r.get_json()['data']['id']

    def test_reject_requires_comment(self, app, admin_client, op_client, seed):
        req_id = self._submit(op_client, seed)
        r = admin_client.post(f'/api/v2/devices/export-password-reviews/{req_id}',
                              json={'action': 'reject', 'comment': ''})
        assert r.status_code == 400
        with app.app_context():
            assert db.session.get(DeviceExportRequest, req_id).status == 'pending'

    def test_reject_flow(self, app, admin_client, op_client, seed):
        req_id = self._submit(op_client, seed)
        r = admin_client.post(f'/api/v2/devices/export-password-reviews/{req_id}',
                              json={'action': 'reject', 'comment': '申请理由不充分'})
        assert r.get_json()['code'] == 0
        with app.app_context():
            req = db.session.get(DeviceExportRequest, req_id)
            assert req.status == 'rejected'
            assert req.review_comment == '申请理由不充分'
            assert req.reviewed_by_user_id == User.query.filter_by(username='admin').first().id
            # 申请人收到通知
            op = User.query.filter_by(username='op').first()
            msgs = Notification.query.filter_by(user_id=op.id, category='device').all()
            assert any('已驳回' in m.title for m in msgs)
            # 审计
            assert AuditLog.query.filter_by(action='device:export_review').count() == 1

    def test_approve_generates_encrypted_package(self, app, admin_client, op_client, seed,
                                                 monkeypatch, tmp_path):
        from blueprints import vue_export
        monkeypatch.setattr(vue_export, 'EXPORT_DIR', str(tmp_path))
        req_id = self._submit(op_client, seed)
        r = admin_client.post(f'/api/v2/devices/export-password-reviews/{req_id}',
                              json={'action': 'approve', 'comment': '同意'})
        assert r.get_json()['code'] == 0
        with app.app_context():
            req = db.session.get(DeviceExportRequest, req_id)
            assert req.status == 'approved'
            assert req.file_token
            assert ExportFile.query.filter_by(token=req.file_token).first()
            # 审计
            assert AuditLog.query.filter_by(action='device:export_review').count() == 1

    def test_approval_uses_requester_customer_scope(self, app, admin_client, op_client, seed,
                                                    monkeypatch, tmp_path):
        from blueprints import vue_export
        monkeypatch.setattr(vue_export, 'EXPORT_DIR', str(tmp_path))
        with app.app_context():
            other_customer = Customer(name='scope-hidden-customer')
            db.session.add(other_customer)
            db.session.flush()
            db.session.add(Device(
                customer_id=other_customer.id,
                device_name='scope-hidden-device',
                password_encrypted=encrypt_password('must-not-export'),
            ))
            db.session.commit()

        r = op_client.post('/api/v2/devices/export-password-request', json={
            'filters': {'preset': 'password'},
            'reason': 'scope regression',
        })
        req_id = r.get_json()['data']['id']
        r = admin_client.post(f'/api/v2/devices/export-password-reviews/{req_id}',
                              json={'action': 'approve', 'comment': 'ok'})
        assert r.get_json()['code'] == 0

        with app.app_context():
            audit = AuditLog.query.filter_by(action='device:export_review').one()
            assert '1 ' in audit.detail

    def test_approve_then_download_one_time_with_password(self, app, admin_client, op_client,
                                                          seed, monkeypatch, tmp_path):
        from datetime import date
        from blueprints import vue_export
        monkeypatch.setattr(vue_export, 'EXPORT_DIR', str(tmp_path))
        req_id = self._submit(op_client, seed)
        admin_client.post(f'/api/v2/devices/export-password-reviews/{req_id}',
                          json={'action': 'approve', 'comment': '同意'})
        with app.app_context():
            token = db.session.get(DeviceExportRequest, req_id).file_token
        # 申请人下载：X-Export-Password 头 + X-Export-Filename 头 + pyzipper 解密验证内容
        r = op_client.get(f'/api/v2/devices/export-password-download/{token}')
        assert r.status_code == 200
        pwd = r.headers.get('X-Export-Password')
        assert pwd
        fname = f'密码客户_设备密码表_{date.today().isoformat()}.xlsx'
        assert r.headers.get('X-Export-Filename') == fname
        import pyzipper
        zf = pyzipper.AESZipFile(io.BytesIO(r.data))
        zf.setpassword(pwd.encode())
        names = zf.namelist()
        assert names == [fname]
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(fname)), read_only=True)
        ws = wb.active
        header = [c.value for c in ws[1]]
        assert '登录密码' in header
        rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        row = dict(zip(header, rows[0]))
        assert row['名称'] == '密码设备'
        assert row['登录密码'] == 'p@ssw0rd!'
        # 一次性：二次下载 404
        r2 = op_client.get(f'/api/v2/devices/export-password-download/{token}')
        assert r2.status_code == 404
        with app.app_context():
            assert db.session.get(DeviceExportRequest, req_id).downloaded_at is not None
            assert AuditLog.query.filter_by(action='device:export_download').count() == 1

    def test_other_user_cannot_download(self, app, admin_client, op_client, viewer_client,
                                        seed, monkeypatch, tmp_path):
        from blueprints import vue_export
        monkeypatch.setattr(vue_export, 'EXPORT_DIR', str(tmp_path))
        req_id = self._submit(op_client, seed)
        admin_client.post(f'/api/v2/devices/export-password-reviews/{req_id}',
                          json={'action': 'approve', 'comment': '同意'})
        with app.app_context():
            token = db.session.get(DeviceExportRequest, req_id).file_token
        r = viewer_client.get(f'/api/v2/devices/export-password-download/{token}')
        assert r.status_code == 403

    def test_already_reviewed_rejected(self, app, admin_client, op_client, seed):
        req_id = self._submit(op_client, seed)
        admin_client.post(f'/api/v2/devices/export-password-reviews/{req_id}',
                          json={'action': 'reject', 'comment': '驳回'})
        r = admin_client.post(f'/api/v2/devices/export-password-reviews/{req_id}',
                              json={'action': 'approve', 'comment': 'x'})
        assert r.status_code == 400

    def test_unknown_token_404(self, op_client):
        r = op_client.get('/api/v2/devices/export-password-download/not_a_token')
        assert r.status_code == 404
