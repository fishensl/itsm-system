# -*- coding: utf-8 -*-
"""W0-S1 设备明文密码收敛验证：
- 设备 JSON 不携带明文密码
- reveal 端点权限门禁 + 401/403 JSON
- Excel 导出密码列按权限收敛
"""
import io

import openpyxl
import pytest

from models import db, Customer, Device, PasswordHistory
from utils.crypto import encrypt_password

PLAIN_PWD = 'S3cret!密码'
OLD_PWD = 'OldPass#2020'


@pytest.fixture()
def device(app):
    """创建带密码设备 + 一条历史密码，返回 (device_id, history_id)"""
    with app.app_context():
        c = Customer(name='密码测试客户')
        db.session.add(c)
        db.session.flush()
        d = Device(customer_id=c.id, device_name='核心交换机SW1',
                   password_encrypted=encrypt_password(PLAIN_PWD))
        db.session.add(d)
        db.session.flush()
        h = PasswordHistory(
            device_id=d.id, password_encrypted=encrypt_password(OLD_PWD),
            changed_by='admin', remark='首次入库')
        db.session.add(h)
        db.session.commit()
        yield d.id, h.id


class TestDeviceJsonNoPassword:
    def test_device_json_excludes_password(self, op_client, device):
        device_id, _ = device
        r = op_client.get(f'/api/devices/{device_id}')
        assert r.status_code == 200
        body = r.get_json()
        assert 'password' not in body
        assert body['has_password'] is True

    def test_device_json_requires_login(self, client, device):
        device_id, _ = device
        r = client.get(f'/api/devices/{device_id}')
        assert r.status_code == 401
        assert r.is_json


class TestRevealEndpoint:
    def test_anonymous_gets_401(self, client, device):
        device_id, _ = device
        r = client.post(f'/api/devices/{device_id}/reveal-password')
        assert r.status_code == 401
        assert r.is_json

    def test_viewer_gets_403_json(self, viewer_client, device):
        """viewer 有 device:view 但无 device:reveal"""
        device_id, _ = device
        r = viewer_client.post(f'/api/devices/{device_id}/reveal-password')
        assert r.status_code == 403
        assert r.is_json

    def test_operator_reveals_current_password(self, op_client, device):
        device_id, _ = device
        r = op_client.post(f'/api/devices/{device_id}/reveal-password')
        assert r.status_code == 200
        assert r.get_json()['password'] == PLAIN_PWD

    def test_operator_reveals_history_password(self, op_client, device):
        device_id, history_id = device
        r = op_client.post(f'/api/devices/{device_id}/reveal-password',
                           data={'history_id': history_id})
        assert r.status_code == 200
        assert r.get_json()['password'] == OLD_PWD

    def test_reveal_history_wrong_device_404(self, op_client, device):
        device_id, _ = device
        r = op_client.post(f'/api/devices/{device_id}/reveal-password',
                           data={'history_id': 999999})
        assert r.status_code == 404


class TestPasswordHistoryApi:
    def test_history_list_has_no_plaintext(self, op_client, device):
        device_id, _ = device
        r = op_client.get(f'/api/devices/{device_id}/password-history')
        assert r.status_code == 200
        rows = r.get_json()
        assert len(rows) == 1
        assert 'password' not in rows[0]
        assert rows[0]['changed_by'] == 'admin'


class TestExportPasswordColumn:
    def _export_headers(self, client, cols):
        r = client.post('/devices/export', data={'export_columns': cols})
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb.active
        return [c.value for c in ws[1]], [[c.value for c in row] for row in ws.iter_rows(min_row=2)]

    def test_viewer_export_strips_password_column(self, viewer_client, device):
        """viewer 无 reveal 权限：即使显式勾选密码列也被剥离"""
        headers, rows = self._export_headers(
            viewer_client, ['device_name', 'password'])
        assert '登录密码' not in headers
        assert '设备名称' in headers
        for row in rows:
            assert PLAIN_PWD not in [str(v) for v in row]

    def test_operator_export_includes_password_with_audit(self, op_client, device):
        headers, rows = self._export_headers(
            op_client, ['device_name', 'password'])
        assert '登录密码' in headers
        pwd_idx = headers.index('登录密码')
        assert rows[0][pwd_idx] == PLAIN_PWD


class TestV2ExportPasswordColumn:
    """V24：v2 导出不再支持密码列直出 —— 含密码列 400，密码只走审核流"""

    def test_v2_export_without_password_ok(self, op_client, device):
        r = op_client.post('/api/v2/devices/export', json={'columns': ['name', 'sn']})
        assert r.get_json()['code'] == 0
        data = r.get_json()['data']
        assert data['filename'].endswith('.xlsx')
        assert data['content']  # base64

    def test_v2_export_password_column_rejected_400(self, op_client, device):
        """即使有 device:reveal 权限，v2 也不允许直出密码列"""
        r = op_client.post('/api/v2/devices/export', json={'columns': ['name', 'password']})
        assert r.status_code == 400
        assert '审核流程' in r.get_json()['message']

    def test_v2_export_password_preset_rejected(self, op_client, device):
        r = op_client.post('/api/v2/devices/export', json={'preset': 'password'})
        assert r.status_code == 400
        assert '审核流程' in r.get_json()['message']

    def test_v2_asset_preset_has_no_password_column(self, op_client, device):
        import base64
        r = op_client.post('/api/v2/devices/export', json={'preset': 'asset'})
        data = r.get_json()['data']
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(data['content'])), read_only=True)
        header = [c.value for c in wb.active[1]]
        assert '登录密码' not in header

    def test_v2_plaintext_only_via_approval_flow(self, app, op_client, admin_client,
                                                 device, monkeypatch, tmp_path):
        """明文密码仅经审核流加密包下发（解密后可见明文）"""
        from blueprints import vue_export
        monkeypatch.setattr(vue_export, 'EXPORT_DIR', str(tmp_path))
        device_id, _ = device
        r = op_client.post('/api/v2/devices/export-password-request', json={
            'filters': {'preset': 'password'}, 'reason': '安全审计'})
        assert r.get_json()['code'] == 0
        req_id = r.get_json()['data']['id']
        r = admin_client.post(f'/api/v2/devices/export-password-reviews/{req_id}',
                              json={'action': 'approve', 'comment': '同意'})
        assert r.get_json()['code'] == 0
        with app.app_context():
            from models import DeviceExportRequest
            token = db.session.get(DeviceExportRequest, req_id).file_token
        r = op_client.get(f'/api/v2/devices/export-password-download/{token}')
        assert r.status_code == 200
        pwd = r.headers.get('X-Export-Password')
        assert pwd
        import pyzipper
        zf = pyzipper.AESZipFile(io.BytesIO(r.data))
        zf.setpassword(pwd.encode())
        wb = openpyxl.load_workbook(io.BytesIO(zf.read('设备密码表.xlsx')), read_only=True)
        header = [c.value for c in wb.active[1]]
        rows = [[c.value for c in row] for row in wb.active.iter_rows(min_row=2)]
        row = dict(zip(header, rows[0]))
        assert row['登录密码'] == PLAIN_PWD
