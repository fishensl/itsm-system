# -*- coding: utf-8 -*-
"""V24 导出筛选：设备三预设列顺序、密码列拒绝、各模块 columns 动态渲染与筛选、空结果"""
import base64
import io
import pytest

from models import db, Customer, Device, Rack, RackInstall, PasswordHistory, Inspection, Ticket, Fault
from utils.crypto import encrypt_password


@pytest.fixture()
def seed(app):
    with app.app_context():
        c = Customer(name='导出客户A')
        c2 = Customer(name='导出客户B')
        db.session.add_all([c, c2])
        db.session.flush()
        d1 = Device(
            customer_id=c.id, device_name='核心交换机', device_type='网络设备',
            brand='华为', model='S5720', serial_number='SN001', ip_address='10.0.0.1',
            port=22, username='admin', password_encrypted=encrypt_password('secret123'),
            login_method='SSH', os_version='V200R019', rule_version='V2.0',
            license_start=None, license_expiry=None, build_date=None,
            is_maintenance=True, is_in_use=True, remark='备注一', location='机房A')
        d2 = Device(customer_id=c2.id, device_name='防火墙', device_type='安全设备',
                    brand='深信服', model='AF-1000', serial_number='SN002',
                    ip_address='10.0.0.2', is_in_use=False, remark='备注二')
        db.session.add_all([d1, d2])
        db.session.flush()
        rack = Rack(customer_id=c.id, name='A-01', location='2F 机房 B 区')
        db.session.add(rack)
        db.session.flush()
        db.session.add(RackInstall(rack_id=rack.id, device_id=d1.id, start_u=3, occupy_u=2))
        db.session.add(PasswordHistory(device_id=d1.id, changed_by='op', remark='改密',
                                       password_encrypted=encrypt_password('oldpwd')))
        i1 = Inspection(title='导出巡检', customer_id=c.id, inspection_date=None,
                        inspector_name='op', overall_status='正常', review_status='已通过')
        t1 = Ticket(number='WO-TEST-001', title='导出工单', customer_id=c.id, status='已完成',
                    priority='中', assigned_to='op', created_by='admin')
        f1 = Fault(title='导出故障', customer_id=c.id, handler='op', fault_type='硬件故障',
                   result='已解决')
        db.session.add_all([i1, t1, f1])
        db.session.commit()
        yield {'c': c.id, 'c2': c2.id, 'd1': d1.id, 'd2': d2.id, 'i1': i1.id,
               't1': t1.id, 'f1': f1.id}


def _decode_xlsx(resp):
    data = resp.get_json()
    assert data['code'] == 0
    raw = base64.b64decode(data['data']['content'])
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    return header, rows


class TestDevicePresets:
    def test_preset_asset_columns(self, op_client, seed):
        r = op_client.post('/api/v2/devices/export', json={'preset': 'asset'})
        header, rows = _decode_xlsx(r)
        assert header == ['客户', '机房位置', '机柜号', '安装位置', '名称', '类型', '品牌', '型号',
                          '序列号', 'IP', '建设时间', '是否维修', '是否在用', '备注']
        assert '登录密码' not in header
        assert len(rows) == 2
        by_name = {dict(zip(header, row))['名称']: dict(zip(header, row)) for row in rows}
        row1 = by_name['核心交换机']
        assert row1['客户'] == '导出客户A'
        assert row1['机房位置'] == '2F 机房 B 区'  # 机柜 Rack.location
        assert row1['机柜号'] == 'A-01'
        assert row1['安装位置'] == '机房A'  # 设备「安装位置」字段 Device.location
        assert row1['是否维修'] == '是'

    def test_preset_password_columns(self, op_client, seed):
        """密码表预设列顺序 + 密码列被 400 拒绝（走申请流）"""
        from blueprints.vue_export import DEVICE_PRESETS
        r = op_client.post('/api/v2/devices/export', json={'preset': 'password'})
        assert r.status_code == 400
        assert '审核流程' in r.get_json()['message']
        # 预设定义本身含密码列且顺序正确
        cols = DEVICE_PRESETS['password']
        assert cols == ['customer', 'rack_location', 'rack_name', 'location', 'name', 'type',
                        'brand', 'model', 'sn', 'ip', 'port', 'login_method', 'username',
                        'password', 'is_in_use', 'pwd_changed_by', 'pwd_changed_at', 'remark']

    def test_preset_version_columns(self, op_client, seed):
        from blueprints.vue_export import DEVICE_PRESETS
        assert DEVICE_PRESETS['version'] == [
            'customer', 'rack_location', 'rack_name', 'location', 'name', 'type', 'brand',
            'model', 'sn', 'ip', 'build_date', 'os_version', 'rule_version', 'license_start',
            'license_expiry', 'is_in_use', 'remark']

    def test_custom_columns_no_password(self, op_client, seed):
        r = op_client.post('/api/v2/devices/export', json={
            'columns': ['name', 'sn', 'pwd_changed_by', 'pwd_changed_at', 'remark']})
        header, rows = _decode_xlsx(r)
        assert header == ['名称', '序列号', '上次修改密码账号', '上次修改密码时间', '备注']
        by_name = {dict(zip(header, row))['名称']: dict(zip(header, row)) for row in rows}
        row1 = by_name['核心交换机']
        assert row1['上次修改密码账号'] == 'op'
        assert row1['上次修改密码时间']

    def test_custom_columns_with_password_rejected(self, op_client, seed):
        r = op_client.post('/api/v2/devices/export', json={
            'columns': ['name', 'password']})
        assert r.status_code == 400
        assert '审核流程' in r.get_json()['message']

    def test_unknown_column_400(self, op_client, seed):
        r = op_client.post('/api/v2/devices/export', json={'columns': ['name', 'nope']})
        assert r.status_code == 400

    def test_customer_filter(self, op_client, seed):
        r = op_client.post('/api/v2/devices/export', json={
            'columns': ['name'], 'customer_id': seed['c2']})
        _, rows = _decode_xlsx(r)
        assert [row[0] for row in rows] == ['防火墙']

    def test_password_column_needs_login_flow_not_reveal_perm(self, op_client, seed):
        """device:reveal 权限也不允许 v2 直出密码（统一走申请流）"""
        r = op_client.post('/api/v2/devices/export', json={'columns': ['name', 'password']})
        assert r.status_code == 400


class TestModuleExports:
    def test_inspection_export_columns_and_filter(self, op_client, seed):
        r = op_client.post('/api/inspections/export', json={
            'columns': ['title', 'customer', 'review_status'],
            'customer_id': seed['c']})
        header, rows = _decode_xlsx(r)
        assert header == ['标题', '客户', '审核状态']
        assert rows[0][:2] == ['导出巡检', '导出客户A']

    def test_inspection_export_date_filter(self, op_client, seed):
        r = op_client.post('/api/inspections/export', json={
            'columns': ['title'],
            'date_from': '2099-01-01', 'date_to': '2099-12-31'})
        _, rows = _decode_xlsx(r)
        assert rows == []

    def test_inspection_bundle_empty_400(self, op_client, seed):
        r = op_client.post('/api/inspections/export-bundle', json={
            'items': ['report'], 'date_from': '2099-01-01'})
        assert r.status_code == 400

    def test_inspection_bundle_invalid_item_400(self, op_client, seed):
        r = op_client.post('/api/inspections/export-bundle', json={'items': ['nope']})
        assert r.status_code == 400

    def test_inspection_bundle_no_items_400(self, op_client, seed):
        r = op_client.post('/api/inspections/export-bundle', json={'items': []})
        assert r.status_code == 400

    def test_ticket_export(self, op_client, seed):
        r = op_client.post('/api/tickets/export', json={
            'columns': ['number', 'title', 'status', 'assigned_to']})
        header, rows = _decode_xlsx(r)
        assert header == ['工单号', '标题', '状态', '处理人']
        assert rows[0][1:] == ['导出工单', '已完成', 'op']

    def test_fault_export(self, op_client, seed):
        r = op_client.post('/api/faults/export', json={
            'columns': ['title', 'fault_type', 'result']})
        header, rows = _decode_xlsx(r)
        assert header == ['标题', '故障类型', '处理结果']
        assert rows[0] == ['导出故障', '硬件故障', '已解决']

    def test_customer_export_date_filter(self, admin_client, seed):
        """客户导出需 customer:export（admin 拥有）"""
        r = admin_client.post('/api/v2/customers/export', json={
            'columns': ['name', 'level'],
            'date_from': '2099-01-01'})
        _, rows = _decode_xlsx(r)
        assert rows == []

    def test_customer_export_columns(self, admin_client, seed):
        r = admin_client.post('/api/v2/customers/export', json={'columns': ['name', 'created_at']})
        header, rows = _decode_xlsx(r)
        assert header == ['客户名称', '创建时间']
        names = {row[0] for row in rows}
        assert '导出客户A' in names

    def test_spare_export(self, app, op_client, seed):
        from models import SparePart, SpareStock
        with app.app_context():
            p = SparePart(name='网线', code='WL-001', category='线缆', unit='根')
            db.session.add(p)
            db.session.flush()
            db.session.add(SpareStock(spare_part_id=p.id, quantity=10, location='库房A'))
            db.session.commit()
        r = op_client.post('/api/spare-parts/export', json={
            'columns': ['code', 'name', 'quantity']})
        header, rows = _decode_xlsx(r)
        assert header == ['备件编码', '名称', '库存数量']
        assert ['WL-001', '网线', 10] in rows

    def test_export_requires_permission(self, viewer_client, seed):
        r = viewer_client.post('/api/faults/export', json={'columns': ['title']})
        assert r.status_code == 200  # viewer 有 fault:view
        r = viewer_client.post('/api/spare-parts/export', json={'columns': ['name']})
        assert r.status_code == 200  # viewer 有 spare:view
