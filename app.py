#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ITSM 简易运维管理系统 - 主应用"""

import json
import os
from datetime import datetime, date, timedelta

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, send_from_directory, jsonify, abort)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.middleware.proxy_fix import ProxyFix

import json

from models import db, User, Customer, Device, Inspection, Fault, Inspector, DeviceType, PasswordHistory, FaultType, InspectionTemplate
from models import Region, Ticket, KnowledgeBase, Brand, NetworkType, CustomField, DeviceCredential, DeviceInterface, InspectionTask, TicketLog
from models import SparePart, SpareStock, PurchaseOrder, SalesOrder
from models import Opportunity, Quotation, Contract, Project
from models import AIConfig, Topology, DeviceConfigBackup, DeviceCollectTask
from models import Department, CustomerCategory, FormDraft, DeviceSubType, UserDashboardPreference, UserPermission, Permission
from models import Role, RolePermission
from models import InspectionDeviceTemplate, InspectionTaskTemplate
from utils.crypto import encrypt_password, decrypt_password
from utils.report_generator import generate_inspection_report, generate_fault_report
from utils.pagination import paginate, paginate_render_args
from utils.permission import require_permission, get_user_permissions, register_template_functions, FAULT_CATEGORY_LEVEL1, FAULT_CATEGORY_LEVEL2, ROOT_CAUSE_CATEGORIES, SEVERITY_LEVELS, is_supervisor
from utils.upload import validate_upload, save_temp_upload, open_excel, cleanup_temp_file, ALLOWED_EXCEL_EXT, MAX_IMPORT_ROWS
from config import Config, setup_logging, setup_security_headers

app = Flask(__name__)
# 经 nginx/反代时识别 X-Forwarded-Proto，使 request.is_secure 正确反映外部协议
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
app.config['SECRET_KEY'] = Config.SECRET_KEY
app.config['SQLALCHEMY_DATABASE_URI'] = Config.SQLALCHEMY_DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = Config.SQLALCHEMY_TRACK_MODIFICATIONS
app.config['MAX_CONTENT_LENGTH'] = Config.MAX_CONTENT_LENGTH
# CSRF：默认对所有 POST/PUT/PATCH/DELETE 启用
app.config['WTF_CSRF_HEADERS'] = ['X-CSRFToken', 'X-CSRF-Token']
app.config['WTF_CSRF_TIME_LIMIT'] = 60 * 60 * 4  # 4 小时

# Limiter：基于 IP 的限流（限流存储使用内存，单进程足够）
app.config['RATELIMIT_STORAGE_URI'] = 'memory://'
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=[],  # 默认不限；个别路由显式声明
    headers_enabled=True,
)

setup_logging(app)
register_template_functions(app)
setup_security_headers(app)

# V13: 证书选项注入 Jinja 全局，模板按分组渲染 checkbox
from utils.cert_options import CERT_CATEGORIES as _CERT_CATEGORIES
app.jinja_env.globals['CERT_CATEGORIES'] = _CERT_CATEGORIES

# 读取应用版本号
_VERSION_FILE = os.path.join(os.path.dirname(__file__), 'VERSION')
try:
    with open(_VERSION_FILE, 'r', encoding='utf-8') as _vf:
        _APP_VERSION = _vf.read().strip()
except Exception:
    _APP_VERSION = 'unknown'
app.jinja_env.globals['APP_VERSION'] = _APP_VERSION

# CSRF 必须在 register_blueprints 之前 init，但 login 路由要豁免（外部 POST）
csrf = CSRFProtect(app)


# CSRF token 同步写入非 HttpOnly cookie，供前端 JS 读取
@app.after_request
def _set_csrf_cookie(response):
    try:
        # 触发 token 生成（写入 session + g.csrf_token）
        token = generate_csrf()
        response.set_cookie(
            'csrf_token', token,
            max_age=60 * 60 * 4,
            httponly=False,  # 允许 JS 读取以放进 X-CSRFToken 头
            samesite='Lax',
            secure=request.is_secure,  # 跟随真实请求协议：HTTPS 才带 Secure，LAN 走 HTTP 也能落地
        )
    except Exception:
        pass
    return response

db.init_app(app)

# flask-migrate（Alembic）接管 schema 演进：替代旧 utils/seed_permissions.py 的 PRAGMA 自动 ADD COLUMN
# init_db() 内部会调 flask db upgrade 应用 migrations/ 下的迁移脚本
from flask_migrate import Migrate
migrate = Migrate(app, db)

# 注册新增蓝图模块
from blueprints import register_blueprints
register_blueprints(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'


# ==================== 全局错误处理 ====================
@app.errorhandler(404)
def err_404(e):
    return render_template('errors/error.html', code=404,
                           title='页面未找到', message='您访问的页面不存在或已被移除。',
                           show_back=True), 404


@app.errorhandler(500)
def err_500(e):
    app.logger.exception('500 错误: %s', e)
    return render_template('errors/error.html', code=500,
                           title='服务器内部错误', message='抱歉，服务器处理您的请求时出错。请稍后重试或联系管理员。',
                           show_back=True), 500


@app.errorhandler(403)
def err_403(e):
    return render_template('errors/error.html', code=403,
                           title='权限不足', message='您没有权限访问此页面。',
                           show_back=True), 403


@app.errorhandler(413)
def err_413(e):
    return render_template('errors/error.html', code=413,
                           title='文件过大', message='上传的文件超过系统允许的大小限制（默认 100MB）。',
                           show_back=True), 413


# 豁免 CSRF 的 API 装饰器：用于 /api/* 端点（这些端点通常基于 session/auth 鉴权，
# 不暴露给第三方站点。CSRF 豁免是为了避免 fetch 调用被拒。）
def api_view(func):
    """标记为 API 端点：自动豁免 CSRF"""
    return csrf.exempt(func)


# 注入 csrf_token() 到所有模板（也可用 {{ csrf_token() }} 直接调用）
@app.context_processor
def inject_csrf_token():
    return {'csrf_token': generate_csrf}


# 注入侧栏配置到所有模板
@app.context_processor
def inject_sidebar():
    """每个请求渲染时，根据当前用户的偏好返回侧栏分组"""
    from utils.sidebar_config import get_user_sidebar_groups
    try:
        from flask_login import current_user
        if current_user.is_authenticated:
            groups = get_user_sidebar_groups(current_user)
        else:
            from utils.sidebar_config import get_default_groups
            groups = [
                {
                    'key': g['key'],
                    'title': g['title'],
                    'icon': g['icon'],
                    'enabled': True,
                    'single_link': g.get('single_link'),
                    'children': g.get('children', []),
                }
                for g in get_default_groups()
            ]
    except Exception:
        from utils.sidebar_config import get_default_groups
        groups = [
            {
                'key': g['key'],
                'title': g['title'],
                'icon': g['icon'],
                'enabled': True,
                'single_link': g.get('single_link'),
                'children': g.get('children', []),
            }
            for g in get_default_groups()
        ]
    return {'sidebar_groups': groups, 'request_path': request.path}


@app.template_filter('from_json')
def from_json_filter(value):
    try:
        return json.loads(value) if value else []
    except:
        return []


@login_manager.user_loader
def load_user(user_id):
    # 仅加载启用账号：停用用户的现有 session 立即失效
    return User.query.filter_by(id=int(user_id), is_active=True).first()


# ---------- 首页 ----------
@app.route('/')
@login_required
def index():
    from sqlalchemy import func, or_

    # 强制刷新缓存，确保统计为最新
    db.session.expire_all()

    me = current_user
    role = me.role or 'viewer'
    me_realname = me.realname or me.username

    # ---- 各模块计数 ----
    counts = {
        'customer': Customer.query.count(),
        'device': Device.query.count(),
        'device_in_use': Device.query.filter_by(is_in_use=True).count(),
        'inspection': Inspection.query.count(),
        'inspection_task': InspectionTask.query.count(),
        'inspection_pending': InspectionTask.query.filter(InspectionTask.status.in_(['待执行', '执行中'])).count(),
        'fault': Fault.query.count(),
        'fault_pending': Fault.query.filter(Fault.result != '已解决').count(),
        'ticket': Ticket.query.count(),
        'ticket_pending': Ticket.query.filter(~Ticket.status.in_(['已验收', '已关闭'])).count(),
        'kb': KnowledgeBase.query.count(),
        'spare': SparePart.query.count(),
        'region': Region.query.count(),
        'opp': Opportunity.query.count(),
        'opp_deal': Opportunity.query.filter_by(stage='成交').count(),
        'quote': Quotation.query.count(),
        'contract': Contract.query.count(),
        'project': Project.query.count(),
        'project_active': Project.query.filter_by(status='进行中').count(),
    }

    # 库存预警：用单次聚合查询替换逐项 N+1
    stock_total = db.session.query(
        SparePart.id, SparePart.min_stock,
        func.coalesce(func.sum(SpareStock.quantity), 0)
    ).outerjoin(SpareStock, SpareStock.spare_part_id == SparePart.id
    ).group_by(SparePart.id).all()
    counts['stock_alerts'] = sum(
        1 for _spid, min_s, qty in stock_total if (min_s or 0) > 0 and qty < min_s
    )

    # 预加载所有客户字典（消除 N+1 查询）
    customer_map = {c.id: c.name for c in Customer.query.all()}

    # ---- 即将到期授权（仅 admin / operator 可看）----
    today = date.today()
    deadline = today + timedelta(days=30)
    expiring_devices_data = []
    if role in ('admin', 'operator', 'viewer'):
        expiring_devices = Device.query.filter(
            Device.license_expiry.isnot(None),
            Device.license_expiry <= deadline
        ).order_by(Device.license_expiry).limit(8).all()
        for d in expiring_devices:
            remaining = (d.license_expiry - today).days if d.license_expiry else 0
            expiring_devices_data.append({
                'id': d.id, 'device_name': d.device_name,
                'customer_name': customer_map.get(d.customer_id, '-'),
                'license_expiry': d.license_expiry.strftime('%Y-%m-%d') if d.license_expiry else '',
                'remaining_days': remaining,
            })

    # ---- 我的待处理任务 ----
    my_tasks = []
    my_tickets = []
    my_insp_tasks = []
    my_faults = []

    if role in ('admin', 'operator'):
        my_tickets = Ticket.query.filter(
            Ticket.assigned_to.in_([me_realname, me.username]),
            ~Ticket.status.in_(['已验收', '已关闭'])
        ).order_by(Ticket.created_at.desc()).limit(8).all()
        for t in my_tickets:
            my_tasks.append({
                'type_label': '工单', 'type_color': 'danger',
                'title': t.title,
                'sub': f"{customer_map.get(t.customer_id, '-')} · {t.priority} · {t.status}",
                'url': f"/tickets/{t.id}",
                'time': t.created_at.strftime('%m-%d %H:%M') if t.created_at else '',
            })

        # 巡检任务：通过关联 user_id 查 Inspector（V13 后 name 是 property 不是列，不能 filter_by(name=)）
        insp = Inspector.query.filter_by(user_id=me.id).first()
        my_iid = str(insp.id) if insp else None
        if my_iid:
            for t in InspectionTask.query.order_by(InspectionTask.id.desc()).limit(50).all():
                if my_iid in (t.inspector_ids or '').split(',') and t.status in ('待执行', '执行中'):
                    my_insp_tasks.append(t)
            my_insp_tasks = my_insp_tasks[:5]
        for t in my_insp_tasks:
            my_tasks.append({
                'type_label': '巡检', 'type_color': 'primary',
                'title': t.title,
                'sub': f"{customer_map.get(t.customer_id, '-')} · {t.status} · {t.task_type}",
                'url': f"/task-schedule/{t.id}",
                'time': (t.planned_start.strftime('%m-%d') if t.planned_start else '') + '~' + (t.planned_end.strftime('%m-%d') if t.planned_end else ''),
            })

        # 待处理故障
        my_faults = Fault.query.filter(Fault.result != '已解决').order_by(Fault.fault_time.desc()).limit(5).all()
        for f in my_faults:
            my_tasks.append({
                'type_label': '故障', 'type_color': 'warning',
                'title': f.title,
                'sub': f"{customer_map.get(f.customer_id, '-')} · {f.fault_type or '-'}",
                'url': f"/faults/{f.id}",
                'time': f.fault_time.strftime('%m-%d %H:%M') if f.fault_time else '',
            })

    elif role == 'sales':
        # 销售看待处理商机
        my_opps = Opportunity.query.filter(
            Opportunity.owner.in_([me_realname, me.username]),
            ~Opportunity.stage.in_(['成交', '失败'])
        ).order_by(Opportunity.expected_close_date.asc().nullslast()).limit(8).all()
        for o in my_opps:
            my_tasks.append({
                'type_label': '商机', 'type_color': 'primary',
                'title': o.title,
                'sub': f"{customer_map.get(o.customer_id, '-')} · {o.stage} · {o.expected_amount or 0}",
                'url': "/opportunities",
                'time': o.expected_close_date.strftime('%Y-%m-%d') if o.expected_close_date else '-',
            })
        # 销售也看进行中合同
        my_contracts = Contract.query.filter(Contract.status == '执行中').order_by(Contract.end_date.asc().nullslast()).limit(5).all()
        for c in my_contracts:
            my_tasks.append({
                'type_label': '合同', 'type_color': 'success',
                'title': c.title,
                'sub': f"{customer_map.get(c.customer_id, '-')} · {c.amount or 0}",
                'url': "/contracts",
                'time': c.end_date.strftime('%Y-%m-%d') if c.end_date else '-',
            })

    my_tasks = my_tasks[:8]

    # 最近巡检
    recent_inspections_data = []
    if role in ('admin', 'operator', 'viewer'):
        for i in Inspection.query.order_by(Inspection.id.desc()).limit(5).all():
            recent_inspections_data.append({
                'id': i.id, 'title': i.title,
                'customer_name': customer_map.get(i.customer_id, '-'),
                'inspection_date': i.inspection_date.strftime('%Y-%m-%d') if i.inspection_date else '',
                'overall_status': i.overall_status,
            })

    # 设备类型分布
    device_type_stats = []
    if role in ('admin', 'operator', 'viewer'):
        device_type_stats = db.session.query(
            Device.device_type, func.count(Device.id)
        ).group_by(Device.device_type).order_by(func.count(Device.id).desc()).all()

    # ---- 按角色组装统计卡片（从用户偏好或角色默认读取）----
    def card(label, value, sub, icon, accent):
        return {'label': label, 'value': value, 'sub': sub, 'icon': icon, 'accent': accent}

    # 合并故障+工单统计
    ticket_total = counts['ticket'] + counts['fault']
    ticket_pending_total = counts['ticket_pending'] + counts['fault_pending']

    # 卡片数据工厂
    CARD_VALUES = {
        'customer':      lambda: card('客户总数', counts['customer'], f"{counts['region']} 个地区", 'bi-people', '#2563eb'),
        'device':        lambda: card('设备总数', counts['device'], f"在用 {counts['device_in_use']}", 'bi-router', '#059669'),
        'device_online': lambda: card('在线设备', counts['device_in_use'], '可正常通讯', 'bi-check-circle', '#16a34a'),
        'inspection':    lambda: card('巡检记录', counts['inspection'], f"{counts['inspection_pending']} 个待执行任务" if role in ('admin','operator') else f"任务 {counts['inspection_task']}", 'bi-clipboard-check', '#7c3aed'),
        'ticket':        lambda: card('工单总数', ticket_total, f"{ticket_pending_total} 待处理（含遗留故障）", 'bi-ticket-detailed', '#f59e0b'),
        'kb':            lambda: card('知识条目', counts['kb'], '故障案例与手册' if role in ('admin','viewer') else '快速查询', 'bi-book', '#0891b2'),
        'spare':         lambda: card('备件档案' if role!='operator' else '备件预警', counts['spare'] if role!='operator' else counts['stock_alerts'], f"{counts['stock_alerts']} 库存预警" if role!='operator' else f"备件 {counts['spare']} 项", 'bi-archive', '#16a34a' if role!='operator' else '#ea580c'),
        'opp':           lambda: card('商机跟进', counts['opp'], f"{counts['opp_deal']} 成交", 'bi-lightbulb', '#475569'),
        'quote':         lambda: card('报价单', counts['quote'], '草稿/已发送', 'bi-file-earmark-text', '#7c3aed'),
        'contract':      lambda: card('合同总数', counts['contract'], f"{counts['project_active']} 项目执行中", 'bi-file-earmark-lock', '#ea580c'),
        'project':       lambda: card('项目', counts['project'], f"{counts['project_active']} 进行中", 'bi-diagram-3', '#db2777'),
        'stock_alert':   lambda: card('备件预警', counts['stock_alerts'], f"备件 {counts['spare']} 项", 'bi-exclamation-diamond', '#ea580c'),
        'expiring':      lambda: card('即将到期授权', len(expiring_devices_data), '30天内到期', 'bi-shield-exclamation', '#db2777'),
        'my_tasks':      lambda: card('我的待办', len(my_tasks), '工单+巡检' if role in ('admin','operator') else ('商机+合同' if role=='sales' else '待处理'), 'bi-person-check', '#2563eb'),
    }

    # 从偏好或默认生成卡片列表
    preferred_cards = get_dashboard_cards(current_user)
    metrics = []
    for ck in preferred_cards:
        if ck in CARD_VALUES:
            try:
                metrics.append(CARD_VALUES[ck]())
            except Exception:
                pass
    if not metrics:
        # fallback
        metrics = [CARD_VALUES['ticket'](), CARD_VALUES['device'](), CARD_VALUES['customer']()]

    # ---- 快捷入口（按角色）----
    qe = lambda url, title, sub, icon: {'url': url, 'title': title, 'sub': sub, 'icon': icon}
    if role == 'admin':
        quick_entries = [
            qe('/customers', '客户管理', '客户信息维护', 'bi-people'),
            qe('/devices', '设备管理', '设备档案与密码', 'bi-router'),
            qe('/inspections', '巡检管理', '巡检记录与任务', 'bi-clipboard-check'),
            qe('/tickets', '工单管理', '派单/接单/处理', 'bi-ticket-detailed'),
            qe('/knowledge-base', '知识库', '故障案例与手册', 'bi-book'),
            qe('/ai-config', 'AI 对接', '智能巡检与分析', 'bi-robot'),
            qe('/users', '用户管理', '账号与角色', 'bi-people-fill'),
            qe('/dashboard/reports', '数据报表', 'ECharts 可视化', 'bi-bar-chart'),
        ]
    elif role == 'operator':
        quick_entries = [
            qe('/devices', '设备管理', '设备档案与密码', 'bi-router'),
            qe('/tickets', '工单处理', '我的工单', 'bi-ticket-detailed'),
            qe('/task-schedule/', '任务安排', '执行计划巡检', 'bi-tasks'),
            qe('/inspections', '巡检记录', '提交巡检报告', 'bi-clipboard-check'),
            qe('/knowledge-base', '知识库', '快速查询', 'bi-book'),
            qe('/topologies', '拓扑图', '网络结构', 'bi-diagram-3'),
        ]
    elif role == 'sales':
        quick_entries = [
            qe('/customers', '客户管理', '客户信息', 'bi-people'),
            qe('/opportunities', '商机跟进', '阶段推进', 'bi-lightbulb'),
            qe('/quotations', '报价单', '生成报价', 'bi-file-earmark-text'),
            qe('/contracts', '合同管理', '合同执行', 'bi-file-earmark-lock'),
            qe('/projects', '项目管理', '项目进度', 'bi-diagram-3'),
            qe('/tickets', '客户报修', '替客户提单', 'bi-ticket-detailed'),
        ]
    else:
        quick_entries = [
            qe('/customers', '客户管理', '查看客户', 'bi-people'),
            qe('/devices', '设备管理', '查看设备', 'bi-router'),
            qe('/tickets', '工单管理', '查看工单', 'bi-ticket-detailed'),
            qe('/dashboard/reports', '数据报表', '统计分析', 'bi-bar-chart'),
        ]

    # V3: 主管视角数据
    supervisor_data = None
    is_supervisor_user = is_supervisor(current_user)
    if is_supervisor_user and current_user.department_id:
        dept = Department.query.get(current_user.department_id)
        if dept:
            dept_user_ids = [u.id for u in User.query.filter_by(department_id=dept.id, is_active=True).all()]
            dept_tasks = InspectionTask.query.filter(
                or_(
                    InspectionTask.assigned_to_user_id.in_(dept_user_ids),
                    InspectionTask.dispatched_by == current_user.id,
                )
            ).order_by(InspectionTask.planned_start.asc()).all()
            today = date.today()
            qm = (today.month - 1) // 3
            q_start = today.replace(month=qm * 3 + 1, day=1)
            # V18: 本季度双侧夹紧，避免把 Q4/明年任务算进"本季度"（老代码漏了上界）
            q_end_month = qm * 3 + 3
            import calendar as _cal
            q_end = today.replace(
                month=q_end_month, day=_cal.monthrange(today.year, q_end_month)[1])
            q_tasks = [t for t in dept_tasks
                       if t.planned_start and q_start <= t.planned_start <= q_end]
            supervisor_data = {
                'dept_name': dept.name,
                'dept_total': len(dept_tasks),
                'dept_pending': len([t for t in dept_tasks if t.status == '待执行' and not t.assigned_to_user_id]),
                'dept_in_progress': len([t for t in dept_tasks if t.status == '执行中']),
                'dept_completed_q': len([t for t in q_tasks if t.status == '已完成']),
                'dept_completion_rate': f"{len([t for t in q_tasks if t.status == '已完成']) * 100 // max(len(q_tasks), 1)}%",
                'dept_members': [{'name': u.realname or u.username, 'id': u.id,
                                  'task_count': len([t for t in dept_tasks if t.assigned_to_user_id == u.id]),
                                  'completed': len([t for t in dept_tasks if t.assigned_to_user_id == u.id and t.status == '已完成'])}
                                 for u in User.query.filter_by(department_id=dept.id, is_active=True).all()],
                'quarter_start': q_start.strftime('%Y-%m-%d'),
            }

    # V3: 草稿提醒
    draft_list = []
    if current_user.is_authenticated:
        draft_list = FormDraft.query.filter_by(user_id=current_user.id).order_by(FormDraft.updated_at.desc()).limit(5).all()

    return render_template('index.html',
                           counts=counts,
                           role=role,
                           role_label_text=role,
                           metrics=metrics,
                           quick_entries=quick_entries,
                           my_tasks=my_tasks,
                           my_ticket_count=len(my_tickets),
                           my_insp_task_count=len(my_insp_tasks),
                           my_fault_count=len(my_faults),
                           expiring_devices=expiring_devices_data,
                           recent_inspections=recent_inspections_data,
                           device_type_stats=device_type_stats,
                           is_supervisor_user=is_supervisor_user,
                           supervisor_data=supervisor_data,
                           draft_list=draft_list,
                           card_pool=DASHBOARD_CARD_POOL,
                           preferred_cards=preferred_cards,
                           role_default_cards=ROLE_DEFAULT_CARDS.get(role, []))


# ---------- 登录 ----------
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit('5 per minute;30 per hour', methods=['POST'])
@csrf.exempt  # 登录页对未登录用户开放，不能强制 CSRF
def login():
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and not user.is_active:
            flash('该账号已停用，请联系管理员', 'danger')
            app.logger.warning(f'停用账号 [{username}] 尝试登录')
            return render_template('login.html')
        if user and user.check_password(password):
            login_user(user)
            app.logger.info(f'用户 [{username}] 登录成功')
            return redirect(url_for('index'))
        flash('用户名或密码错误', 'danger')
        app.logger.warning(f'用户 [{username}] 登录失败')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    app.logger.info(f'用户 [{current_user.username}] 登出')
    logout_user()
    return redirect(url_for('login'))

# ==================== 简化的 admin 路由（暂留 app.py 后续蓝图化）====================
from models import (AIConfig, User as UserM, UserPermission, Permission, Department)
from sqlalchemy.orm import joinedload
from utils.permission import admin_required

@app.route('/users', methods=['GET', 'POST'])
@login_required
@require_permission('user:view')
@admin_required
def user_list():
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        if username and not UserM.query.filter_by(username=username).first():
            from utils.cert_options import parse_cert_form
            u = UserM.create_with_password(
                username=username,
                password=request.form.get('password', 'changeme'),
                realname=request.form.get('realname', ''),
                role=request.form.get('role', 'viewer'),
                department_id=request.form.get('department_id', type=int),
            )
            # V13: 人员主数据扩展
            u.phone = (request.form.get('phone') or '').strip()
            u.email = (request.form.get('email') or '').strip()
            u.set_cert_list(parse_cert_form(request.form.getlist('certifications')))
            db.session.add(u); db.session.commit()
            flash(f'用户 {username} 已创建', 'success')
        return redirect(url_for('user_list'))
    users = UserM.query.options(joinedload(UserM.department_rel)).order_by(UserM.id).all()
    departments = Department.query.order_by(Department.sort_order).all()
    roles = Role.query.filter_by(is_active=True).order_by(Role.sort_order, Role.id).all()
    return render_template('users/list.html', users=users, departments=departments, roles=roles,
                           page=1, total_pages=1, has_prev=False, has_next=False,
                           prev_page=None, next_page=None,
                           total=len(users), start=1 if users else 0, end=len(users))


@app.route('/users/delete/<int:id>', methods=['POST'])
@login_required
@require_permission('user:delete')
@admin_required
def user_delete(id):
    UserM.query.filter_by(id=id).delete()
    db.session.commit()
    flash('已删除', 'success')
    return redirect(url_for('user_list'))


@app.route('/users/add', methods=['POST'])
@login_required
@require_permission('user:add')
@admin_required
def user_add():
    return redirect(url_for('user_list'))


# ==================== 用户编辑（V6.1.4） ====================
@app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@require_permission('user:edit')
@admin_required
def user_edit(id):
    """编辑用户：用户名（unique）、姓名、角色、密码、状态、部门、电话、邮箱、证书"""
    u = UserM.query.get_or_404(id)
    if request.method == 'POST':
        new_username = (request.form.get('username') or '').strip()
        if not new_username:
            flash('用户名不能为空', 'danger')
            return redirect(url_for('user_edit', id=id))
        # 用户名唯一性校验
        if new_username != u.username:
            if UserM.query.filter_by(username=new_username).first():
                flash(f'用户名 "{new_username}" 已被其他账号占用', 'danger')
                return redirect(url_for('user_edit', id=id))
        from utils.cert_options import parse_cert_form
        pwd = request.form.get('password', '').strip()
        u.username = new_username
        u.realname = (request.form.get('realname') or '').strip()
        u.role = request.form.get('role', u.role or 'operator')
        u.department_id = request.form.get('department_id', type=int)
        u.is_active = bool(request.form.get('is_active'))
        # V13: 人员主数据扩展
        u.phone = (request.form.get('phone') or '').strip()
        u.email = (request.form.get('email') or '').strip()
        u.set_cert_list(parse_cert_form(request.form.getlist('certifications')))
        if pwd:
            u.set_password(pwd)
        db.session.commit()
        flash(f'用户 {u.username} 已更新', 'success')
        return redirect(url_for('user_list'))

    # GET：渲染编辑页（用 modal 模式：直接回 user_list 弹窗）
    departments = Department.query.order_by(Department.sort_order).all()
    roles = Role.query.filter_by(is_active=True).order_by(Role.sort_order, Role.id).all()
    return render_template('users/edit.html', u=u, departments=departments, roles=roles)


# ==================== V13: 管理员重置密码 ====================
@app.route('/users/<int:id>/reset_password', methods=['POST'])
@login_required
@require_permission('user:edit')
@admin_required
def user_reset_password(id):
    """管理员强制重置任意账号密码（无需原密码）"""
    u = UserM.query.get_or_404(id)
    new_pwd = (request.form.get('new_password') or '').strip()
    if len(new_pwd) < 6:
        flash('新密码长度至少 6 位', 'danger')
        return redirect(url_for('user_list'))
    u.set_password(new_pwd)
    db.session.commit()
    app.logger.info(f'管理员 [{current_user.username}] 重置了用户 [{u.username}] 的密码')
    flash(f'用户 {u.username} 的密码已重置', 'success')
    return redirect(url_for('user_list'))


# ==================== V13: 用户自助修改密码 ====================
@app.route('/me/change_password', methods=['GET', 'POST'])
@login_required
@limiter.limit('10 per hour')
def me_change_password():
    """登录用户自助修改密码（需校验原密码）"""
    if request.method == 'POST':
        old_pwd = request.form.get('old_password') or ''
        new_pwd = (request.form.get('new_password') or '').strip()
        confirm = (request.form.get('confirm_password') or '').strip()
        if not current_user.check_password(old_pwd):
            flash('原密码错误', 'danger')
            return redirect(url_for('me_change_password'))
        if len(new_pwd) < 6:
            flash('新密码长度至少 6 位', 'danger')
            return redirect(url_for('me_change_password'))
        if new_pwd != confirm:
            flash('两次输入的新密码不一致', 'danger')
            return redirect(url_for('me_change_password'))
        if new_pwd == old_pwd:
            flash('新密码不能与原密码相同', 'warning')
            return redirect(url_for('me_change_password'))
        current_user.set_password(new_pwd)
        db.session.commit()
        app.logger.info(f'用户 [{current_user.username}] 自助修改了密码')
        flash('密码已修改，请使用新密码重新登录', 'success')
        # 改完强制退出，让用户用新密码登录
        from flask_login import logout_user
        logout_user()
        return redirect(url_for('login'))
    return render_template('auth/change_password.html')


@app.route('/system')
@login_required
def system_settings():
    """系统概览页：业务统计 + 部署系统信息（CPU/内存/磁盘/版本）"""
    from sqlalchemy import func as _func
    import platform as _plat
    import sys as _sys
    stats = {
        'user_count': UserM.query.filter_by(is_active=True).count(),
        'user_total': UserM.query.count(),
        'department_count': Department.query.count(),
        'customer_count': Customer.query.count(),
        'device_count': Device.query.count(),
        'topology_count': Topology.query.count(),
        'inspection_count': Inspection.query.count(),
        'ticket_count': Ticket.query.count(),
    }
    # 最近 5 个登录用户
    recent_users = UserM.query.options(joinedload(UserM.department_rel))\
        .order_by(UserM.id.desc()).limit(5).all()

    # ==================== V6.1.2 部署系统信息 ====================
    sys_info = {
        # 系统版本
        'os_name': _plat.system(),
        'os_release': _plat.release(),
        'os_version': _plat.version(),
        'os_platform': _plat.platform(),
        'machine': _plat.machine(),
        'hostname': _plat.node(),
        # Python / Flask
        'python_version': _plat.python_version(),
        'python_impl': _plat.python_implementation(),
    }
    # 主要组件版本
    components = {}
    for name, mod in [
        ('Flask', 'flask'), ('Flask-Login', 'flask_login'),
        ('Flask-SQLAlchemy', 'flask_sqlalchemy'), ('Flask-WTF', 'flask_wtf'),
        ('Flask-Limiter', 'flask_limiter'), ('SQLAlchemy', 'sqlalchemy'),
        ('Werkzeug', 'werkzeug'), ('Jinja2', 'jinja2'),
        ('python-docx', 'docx'), ('openpyxl', 'openpyxl'),
        ('cryptography', 'cryptography'), ('psutil', 'psutil'),
    ]:
        try:
            m = __import__(mod)
            components[name] = getattr(m, '__version__', '-')
        except Exception:
            components[name] = '未安装'

    # 数据库版本
    db_info = {'engine': '-', 'version': '-', 'path': '-'}
    try:
        uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if uri.startswith('sqlite:///'):
            import sqlite3 as _sqlite3
            db_info['engine'] = 'SQLite'
            db_info['version'] = _sqlite3.sqlite_version
            db_path = uri.replace('sqlite:///', '')
            db_info['path'] = db_path
            # 数据库文件大小
            if os.path.isfile(db_path):
                db_info['size_mb'] = round(os.path.getsize(db_path) / (1024 * 1024), 2)
        elif 'mysql' in uri:
            db_info['engine'] = 'MySQL'
            try:
                with db.engine.connect() as conn:
                    r = conn.execute(db.text('SELECT VERSION()')).scalar()
                    db_info['version'] = str(r)
            except Exception:
                pass
        elif 'postgresql' in uri:
            db_info['engine'] = 'PostgreSQL'
            try:
                with db.engine.connect() as conn:
                    r = conn.execute(db.text('SHOW server_version')).scalar()
                    db_info['version'] = str(r)
            except Exception:
                pass
    except Exception as _e:
        app.logger.warning(f'数据库信息获取失败: {_e}')

    # 资源占用（CPU/内存/磁盘）
    resources = {}
    try:
        import psutil as _ps
        cpu_pct = _ps.cpu_percent(interval=0.5)
        cpu_count = _ps.cpu_count(logical=True)
        cpu_count_phy = _ps.cpu_count(logical=False) or cpu_count
        mem = _ps.virtual_memory()
        disk_root = _ps.disk_usage(os.path.abspath(os.sep))
        # 进程信息
        proc = _ps.Process(os.getpid())
        proc_mem = proc.memory_info()
        # 启动时间（系统）
        boot_ts = _ps.boot_time()
        from datetime import datetime as _dt
        boot_str = _dt.fromtimestamp(boot_ts).strftime('%Y-%m-%d %H:%M:%S')
        # 启动时间（应用进程）
        proc_start = _dt.fromtimestamp(proc.create_time()).strftime('%Y-%m-%d %H:%M:%S')

        resources = {
            'cpu_percent': cpu_pct,
            'cpu_count': cpu_count,
            'cpu_count_physical': cpu_count_phy,
            'memory_percent': mem.percent,
            'memory_total_gb': round(mem.total / (1024**3), 2),
            'memory_used_gb': round(mem.used / (1024**3), 2),
            'memory_available_gb': round(mem.available / (1024**3), 2),
            'disk_percent': disk_root.percent,
            'disk_total_gb': round(disk_root.total / (1024**3), 2),
            'disk_used_gb': round(disk_root.used / (1024**3), 2),
            'disk_free_gb': round(disk_root.free / (1024**3), 2),
            'process_memory_mb': round(proc_mem.rss / (1024**2), 2),
            'process_pid': proc.pid,
            'boot_time': boot_str,
            'process_start': proc_start,
            'available': True,
        }
    except Exception as _e:
        app.logger.warning(f'资源占用获取失败: {_e}')
        resources = {'available': False, 'error': str(_e)}

    return render_template('system/index.html',
                           stats=stats,
                           recent_users=recent_users,
                           sys_info=sys_info,
                           components=components,
                           db_info=db_info,
                           resources=resources)


# ==================== 侧栏自定义 ====================
@app.route('/system/sidebar', methods=['GET', 'POST'])
@login_required
@api_view  # POST 路由需要豁免 CSRF（前端用 fetch + JSON body）
def system_sidebar():
    """侧栏自定义页面 / 保存"""
    from utils.sidebar_config import (SIDEBAR_GROUPS, get_user_sidebar_groups, save_user_sidebar)
    if request.method == 'POST':
        # 提交顺序 + 启用/禁用
        payload = request.get_json(silent=True) or {}
        groups_data = payload.get('groups', [])
        if not isinstance(groups_data, list):
            return jsonify({'success': False, 'message': '参数错误'}), 400
        save_user_sidebar(current_user, groups_data)
        return jsonify({'success': True, 'message': '侧栏设置已保存'})
    # GET：渲染编辑页面
    current_groups = get_user_sidebar_groups(current_user)
    return render_template('system/sidebar.html',
                           all_groups=SIDEBAR_GROUPS,
                           current_groups=current_groups)


@app.route('/api/sidebar/reset', methods=['POST'])
@login_required
@api_view
def api_sidebar_reset():
    """重置为默认"""
    from models import UserDashboardPreference, db
    pref = UserDashboardPreference.query.filter_by(user_id=current_user.id).first()
    if pref:
        pref.sidebar_json = None
        db.session.commit()
    return jsonify({'success': True, 'message': '已重置为系统默认'})


@app.route('/permissions')
@login_required
@require_permission('permission:view')
def permission_list():
    """权限管理：展示各角色权限对照（数据来自 DB，自定义角色自动出现）"""
    from utils.permission import PERMISSION_MAP
    # 从 DB 拉所有活跃角色（包含自定义）
    roles = Role.query.filter_by(is_active=True).order_by(Role.sort_order, Role.id).all()
    role_perms = {}
    role_list = []
    # [(code, name, rid), ...] 给模板做"点格子跳到该角色配置"用
    role_meta = {}
    for r in roles:
        perms = frozenset(rp.permission_code for rp in r.role_perms)
        # admin 短路：显示全量
        if r.code == 'admin':
            perms = set(PERMISSION_MAP.keys())
        role_perms[r.code] = list(perms)
        role_list.append((r.code, r.name))
        role_meta[r.code] = r.id
    return render_template('permissions/list.html',
                           role_perms=role_perms,
                           perm_map=PERMISSION_MAP,
                           role_list=role_list,
                           role_meta=role_meta)


@app.route('/ai-config', methods=['GET', 'POST'])
@login_required
def ai_config_page():
    configs = AIConfig.query.order_by(AIConfig.id.desc()).all()
    return render_template('ai_config/list.html', configs=configs)


@app.route('/ai-config/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def ai_config_delete(id):
    AIConfig.query.filter_by(id=id).delete()
    db.session.commit()
    flash('已删除', 'success')
    return redirect(url_for('ai_config_page'))


# /dashboard/reports 路由已删除（与运维管理 /reports 重复，且本路由是空壳）
# 旧链接重定向到统一的 /reports
@app.route('/dashboard/reports')
@login_required
def dashboard_reports():
    return redirect(url_for('ops.report_list'))


@app.route('/exports/download-template/<module>')
@login_required
@require_permission('report:view')
def download_template(module):
    """下载批量导入模板 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1890FF', end_color='096DD9', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    templates = {
        'customer': {
            'name': '客户导入模板',
            'headers': ['客户名称', '联系人', '电话', '邮箱', '所属地区', '地市', '地址',
                        '单位类别', '客户等级',
                        '办公室', '有无驻场', '驻场联系人', '驻场联系方式', '驻场办公室',
                        '有无攻防演练', '巡检频率',
                        '来源', '备注'],
        },
        'device': {
            'name': '设备导入模板',
            'headers': ['所属客户', '设备名称', '设备类型', '品牌', '型号', '序列号', 'IP地址', '端口',
                        '登录用户名', '登录密码', '登录方式', '安装位置', '系统版本',
                        '授权开始日期', '授权截止日期', '规则库版本', '是否维修', '是否在用', '备注'],
        },
        'inspection': {
            'name': '巡检记录导入模板',
            'headers': ['客户名称', '标题', '巡检人员', '巡检日期', '巡检地点', '总体状态', '结论', '备注'],
        },
        'fault': {
            'name': '故障记录导入模板',
            'headers': ['客户名称', '标题', '处理人', '故障时间', '故障类型', '故障描述', '故障原因', '解决方案', '处理结果'],
        },
        'spare': {
            'name': '备件导入模板',
            'headers': ['编码', '名称', '分类', '规格', '单位', '最低库存', '备注'],
        },
        'stock': {
            'name': '库存导入模板',
            'headers': ['备件名称', '位置', '数量', '单价'],
        },
    }

    tpl = templates.get(module)
    if not tpl:
        flash('不支持的导入模板类型', 'danger')
        return redirect(url_for('index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tpl['name']

    for col_idx, h in enumerate(tpl['headers'], 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(h) * 2.5, 18)

    import tempfile
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()

    return send_from_directory(
        os.path.dirname(tmp.name),
        os.path.basename(tmp.name),
        as_attachment=True,
        download_name=f'{tpl["name"]}_{date.today().isoformat()}.xlsx'
    )


# ==================== 客户管理 ====================
@app.route('/customers')
@login_required
@require_permission('customer:view')
def customer_list():
    search = request.args.get('search', '')
    category_id = request.args.get('category_id', type=int)
    page = request.args.get('page', 1, type=int)
    query = Customer.query
    if search:
        query = query.filter(
            Customer.name.contains(search) |
            Customer.contact_person.contains(search) |
            Customer.phone.contains(search)
        )
    if category_id:
        query = query.filter_by(category_id=category_id)
    # 预加载 region_rel 及其 parent，避免列表渲染时 N+1
    from sqlalchemy.orm import joinedload
    query = query.options(joinedload(Customer.region_rel).joinedload(Region.parent))
    query = query.order_by(Customer.id.desc())
    pag = paginate(query, page=page)
    categories = CustomerCategory.query.order_by(CustomerCategory.sort_order).all()
    return render_template('customers/list.html', **paginate_render_args(pag), search=search,
                          categories=categories, current_category_id=category_id or 0)


# ==================== 客户管理（已迁移至 blueprints/customer.py）====================
# ========== 客户定级（保留路由层封装，业务规则见 services/customer_service._calculate_tier）==========
def calculate_customer_tier(device_count, has_onsite, has_drill):
    """根据设备数、驻场、攻防演练自动定级（兼容旧调用；新代码用 customer_service._calculate_tier）"""
    from services.customer_service import _calculate_tier
    return _calculate_tier(device_count, has_onsite, has_drill)

def _sync_customer_device_count(customer_id):
    """同步客户的 device_count 冗余字段"""
    if not customer_id:
        return
    cnt = Device.query.filter_by(customer_id=customer_id).count()
    c = Customer.query.get(customer_id)
    if c:
        c.device_count = cnt
        # 重新定级
        auto_tier = calculate_customer_tier(cnt, c.has_onsite, c.has_drill)
        if c.level not in ('核心', '重点', '常规') or not c.level:
            c.level = auto_tier
        db.session.commit()




# ==================== 设备密码管理 ====================
@app.route('/api/dashboard/opportunity-stages')
@login_required
@api_view
def api_dashboard_opp_stages():
    """商机阶段统计（销售工作台用）"""
    from sqlalchemy import func
    stages = ['初步接触', '需求确认', '方案报价', '商务谈判', '成交', '失败']
    rows = db.session.query(Opportunity.stage, func.count(Opportunity.id))\
        .group_by(Opportunity.stage).all()
    stat = {s: 0 for s in stages}
    for s, c in rows:
        if s in stat: stat[s] = c
    return jsonify({'success': True, 'labels': list(stat.keys()), 'values': list(stat.values())})

# ==================== 工作台偏好 API ====================
# 卡片池定义（所有可选卡片）
DASHBOARD_CARD_POOL = {
    'customer':      {'label': '客户总数',   'icon': 'bi-people',         'accent': '#2563eb'},
    'device':        {'label': '设备总数',   'icon': 'bi-router',         'accent': '#059669'},
    'device_online': {'label': '在线设备',   'icon': 'bi-check-circle',   'accent': '#16a34a'},
    'inspection':    {'label': '巡检记录',   'icon': 'bi-clipboard-check','accent': '#7c3aed'},
    'ticket':        {'label': '工单总数',   'icon': 'bi-ticket-detailed','accent': '#f59e0b'},
    'kb':            {'label': '知识条目',   'icon': 'bi-book',           'accent': '#0891b2'},
    'spare':         {'label': '备件档案',   'icon': 'bi-archive',        'accent': '#16a34a'},
    'opp':           {'label': '商机跟进',   'icon': 'bi-lightbulb',      'accent': '#475569'},
    'quote':         {'label': '报价单',     'icon': 'bi-file-earmark-text','accent': '#7c3aed'},
    'contract':      {'label': '合同总数',   'icon': 'bi-file-earmark-lock','accent': '#ea580c'},
    'project':       {'label': '项目',       'icon': 'bi-diagram-3',      'accent': '#db2777'},
    'stock_alert':   {'label': '备件预警',   'icon': 'bi-exclamation-diamond','accent': '#ea580c'},
    'expiring':      {'label': '授权到期',   'icon': 'bi-shield-exclamation','accent': '#db2777'},
    'my_tasks':      {'label': '我的待办',   'icon': 'bi-person-check',   'accent': '#2563eb'},
}

ROLE_DEFAULT_CARDS = {
    'admin':    ['customer', 'device', 'inspection', 'ticket', 'kb', 'spare', 'opp', 'contract', 'project'],
    'operator': ['device', 'my_tasks', 'ticket', 'inspection', 'kb', 'stock_alert', 'expiring'],
    'sales':    ['customer', 'opp', 'quote', 'contract', 'project', 'my_tasks', 'spare'],
    'viewer':   ['customer', 'device', 'ticket', 'project', 'inspection', 'kb'],
}

def get_dashboard_cards(user):
    """获取用户生效的卡片列表（偏好或角色默认）"""
    pref = UserDashboardPreference.query.filter_by(user_id=user.id).first()
    if pref and pref.cards_json:
        try:
            cards = json.loads(pref.cards_json)
            if isinstance(cards, list) and len(cards) > 0:
                return cards
        except (json.JSONDecodeError, TypeError):
            pass
    return ROLE_DEFAULT_CARDS.get(user.role, ['ticket', 'device', 'customer'])

@app.route('/api/dashboard/preferences')
@login_required
@api_view
def api_dashboard_preferences():
    pref = UserDashboardPreference.query.filter_by(user_id=current_user.id).first()
    cards = get_dashboard_cards(current_user)
    return jsonify({
        'cards': cards,
        'is_custom': pref is not None and pref.cards_json and len(pref.cards_json) > 2,
        'defaults': ROLE_DEFAULT_CARDS.get(current_user.role, []),
        'pool': {k: v for k, v in DASHBOARD_CARD_POOL.items()},
    })

@app.route('/api/dashboard/preferences', methods=['POST'])
@login_required
@api_view
def api_dashboard_preferences_save():
    data = request.get_json(silent=True) or {}
    card_keys = data.get('cards', [])
    # 过滤无效卡片key
    valid = [k for k in card_keys if k in DASHBOARD_CARD_POOL]
    pref = UserDashboardPreference.query.filter_by(user_id=current_user.id).first()
    if not pref:
        pref = UserDashboardPreference(user_id=current_user.id)
        db.session.add(pref)
    pref.cards_json = json.dumps(valid)
    db.session.commit()
    return jsonify({'success': True, 'cards': valid})

@app.route('/api/dashboard/preferences/reset', methods=['POST'])
@login_required
@api_view
def api_dashboard_preferences_reset():
    pref = UserDashboardPreference.query.filter_by(user_id=current_user.id).first()
    if pref:
        db.session.delete(pref)
        db.session.commit()
    cards = ROLE_DEFAULT_CARDS.get(current_user.role, [])
    return jsonify({'success': True, 'cards': cards})


# ---------- 初始化 ----------
def _bootstrap_legacy_db():
    """引导遗留库（由旧 db.create_all + ensure_schema 建好但无 alembic_version）接入 Alembic。

    三种库状态：
      1) 空库：无任何业务表 → 不处理，交给 flask db upgrade 从零建表。
      2) 遗留库：有业务表但无 alembic_version 表 → 其 schema 与 initial_schema 一致
         （interface=VARCHAR(128)、customers.name/tickets.number 无唯一约束），
         故 stamp 到 initial_schema，后续 upgrade 只跑 pg_type_fixes。
      3) 已接入 Alembic：有 alembic_version → 不处理，交给 upgrade。
    返回 True 表示已处理（调用了 stamp），False 表示无需处理。
    """
    from sqlalchemy import inspect as sqla_inspect, text
    insp = sqla_inspect(db.engine)
    all_tables = set(insp.get_table_names())
    if 'alembic_version' in all_tables:
        return False  # 已接入
    business_tables = all_tables - {'alembic_version', 'sqlite_sequence'}
    if not business_tables:
        return False  # 空库，让 upgrade 从零建

    # 遗留库：有业务表但无 alembic_version。先清理可能阻塞 pg_type_fixes 唯一约束的重复数据。
    _dedup_before_unique_constraints()

    # stamp 到 initial_schema（遗留库结构与 initial_schema 一致），之后 upgrade 只需跑 pg_type_fixes
    from flask_migrate import stamp as _migrate_stamp
    import os as _os
    _migrate_stamp(directory=_os.path.join(_os.path.dirname(__file__), 'migrations'),
                   revision='3f82f965fb25')
    print('[INIT] 检测到遗留库（无 alembic_version），已 stamp 到 initial_schema，后续 upgrade 将应用 pg_type_fixes')
    return True


def _dedup_before_unique_constraints():
    """给即将加唯一约束的列清理重复行（保留 id 最小者，其余改名加后缀使其唯一）。

    - customers.name：重名客户给较新者追加 " (重复N)" 后缀
    - tickets.number：重号工单给较新者追加 "-DUP-N" 后缀
    幂等：已是唯一则无操作。失败仅告警不中断（不阻塞启动）。
    """
    from sqlalchemy import text
    try:
        # customers.name
        dup_names = db.session.execute(text(
            "SELECT name, COUNT(*) c FROM customers GROUP BY name HAVING COUNT(*) > 1"
        )).all()
        for name, _cnt in dup_names:
            rows = db.session.execute(text(
                "SELECT id FROM customers WHERE name = :n ORDER BY id"
            ), {'n': name}).all()
            for i, (cid,) in enumerate(rows[1:], start=1):
                new_name = f'{name} (重复{i})'
                # 截断到 128 字符以符合 String(128)
                if len(new_name) > 128:
                    new_name = new_name[:128]
                db.session.execute(text(
                    "UPDATE customers SET name = :nn WHERE id = :id"
                ), {'nn': new_name, 'id': cid})
        # tickets.number
        dup_nums = db.session.execute(text(
            "SELECT number, COUNT(*) c FROM tickets GROUP BY number HAVING COUNT(*) > 1"
        )).all()
        for num, _cnt in dup_nums:
            rows = db.session.execute(text(
                "SELECT id FROM tickets WHERE number = :n ORDER BY id"
            ), {'n': num}).all()
            for i, (tid,) in enumerate(rows[1:], start=1):
                new_num = f'{num}-DUP{i}'
                if len(new_num) > 32:
                    new_num = (num[:32 - len(f'-DUP{i}')] + f'-DUP{i}')
                db.session.execute(text(
                    "UPDATE tickets SET number = :nn WHERE id = :id"
                ), {'nn': new_num, 'id': tid})
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f'[WARN] 去重清理失败（非致命，可能在 pg_type_fixes 加唯一约束时报错）: {e}')


def init_db():
    with app.app_context():
        # Schema 演进交给 Alembic（flask-migrate），替代旧 db.create_all + ensure_schema(PRGAMA)
        from flask_migrate import upgrade as _migrate_upgrade
        import os as _os
        migrations_dir = _os.path.join(_os.path.dirname(__file__), 'migrations')

        # 先引导遗留库（有表但无 alembic_version 的旧 SQLite 库）接入 Alembic
        _bootstrap_legacy_db()

        # 应用所有待执行的迁移（空库会从 initial_schema 一路建到 head；遗留库只跑 pg_type_fixes）
        _migrate_upgrade(directory=migrations_dir)

        # V14: 权限/角色 seed（幂等，仅写数据不改 schema）
        try:
            from utils.seed_permissions import seed_all
            seed_all(app)
        except Exception as e:
            print(f'[WARN] 权限 seed 失败（非致命）: {e}')
            db.session.rollback()

        # 创建默认管理员：仅在系统中不存在任何 admin 角色用户时（首次空库引导），
        # 避免管理员把 admin 改名/删除后，重启又重建 admin/admin123 弱口令后门
        if User.query.filter_by(role='admin').count() == 0:
            admin = User.create_with_password(username='admin', password='admin123', realname='管理员', role='admin')
            db.session.add(admin)
            db.session.commit()
            app.logger.info('[INIT] 默认管理员已创建: admin / admin123')
            print('[OK] 默认用户已创建: admin / admin123')

        # 创建默认设备类型
        if DeviceType.query.count() == 0:
            defaults = ['路由器', '交换机', '防火墙', '服务器', '负载均衡', '无线AP', '光传输', 'UPS电源', '空调', '其他']
            for i, name in enumerate(defaults):
                db.session.add(DeviceType(name=name, sort_order=i))
            db.session.commit()
            print('[OK] 默认设备类型已创建')

        # 创建默认故障类型
        if FaultType.query.count() == 0:
            defaults = ['网络中断', '设备故障', '安全事件', '链路故障', '电源故障', '配置错误', '性能问题', '其他']
            for i, name in enumerate(defaults):
                db.session.add(FaultType(name=name, sort_order=i))
            db.session.commit()
            print('[OK] 默认故障类型已创建')


if __name__ == '__main__':
    init_db()
    print('=' * 50)
    print('=== ITSM 简易运维管理系统 ===')
    print('=' * 50)
    print('默认登录: admin / admin123')
    print('访问地址: http://127.0.0.1:5000')
    print('=' * 50)
    app.run(debug=True, host='127.0.0.1', port=5000)
