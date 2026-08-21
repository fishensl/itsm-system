# -*- coding: utf-8 -*-
"""系统运维端点：schema 修复 / drawio 诊断 / 侧栏保存 / 导入模板下载（SSR 业务页已剥离）"""
import os
from datetime import date
from flask import (request, redirect, url_for,
                   flash, jsonify, current_app)
from flask_login import (login_required, current_user)
from models import db, UserDashboardPreference
from utils.permission import require_permission, admin_required
from utils.compat import deprecated_endpoint


# ==================== 简化的 admin 路由（暂留 app.py 后续蓝图化）====================
@login_required
@admin_required
def repair_schema():
    """只读诊断 DB schema；修复必须通过 Alembic 部署命令执行。"""
    from sqlalchemy import inspect as sqla_inspect, text
    reports = []

    # 关键列及其定义（表名 → (列名, SQL 类型)）— 与 models.py / 迁移保持一致
    # 注：用方言无关的 SQLAlchemy 类型生成补列语句（DATETIME 是 SQLite 专属，
    # PG 上必须用 TIMESTAMP，直接写死 DATETIME 会导致 PG 补列失败）
    CRITICAL_COLUMNS = {
        'inspection_tasks': [
            ('estimated_effort', 'FLOAT'),
            ('actual_effort', 'FLOAT'),
        ],
        'topologies': [
            ('diagram_xml', 'TEXT'),
            ('source', 'VARCHAR(16)'),
            ('thumbnail_path', 'VARCHAR(512)'),
            ('pdf_path', 'VARCHAR(512)'),
            ('vsdx_path', 'VARCHAR(512)'),
            ('updated_at', 'DATETIME'),
        ],
    }

    # 1. 当前 alembic 版本
    try:
        insp = sqla_inspect(db.engine)
        if 'alembic_version' not in (insp.get_table_names()):
            reports.append(('alembic_version', '表不存在（遗留库未接入 Alembic）', 'warn'))
        else:
            ver = db.session.execute(text('SELECT version_num FROM alembic_version')).scalar()
            reports.append(('alembic 当前版本', ver or '(空)', 'info'))
    except Exception as e:
        reports.append(('alembic 查询失败', str(e), 'danger'))

    # 2. 关键列检查 + 缺失则直接 ALTER TABLE 补列
    try:
        insp = sqla_inspect(db.engine)
        existing_tables = set(insp.get_table_names())
        for tbl, cols_def in CRITICAL_COLUMNS.items():
            if tbl not in existing_tables:
                reports.append((f'{tbl} 表', '❌ 表不存在', 'danger'))
                continue
            existing_cols = {c['name'] for c in insp.get_columns(tbl)}
            for col_name, col_type in cols_def:
                if col_name in existing_cols:
                    reports.append((f'{tbl}.{col_name}', '✅ 存在', 'ok'))
                else:
                    reports.append((f'{tbl}.{col_name}',
                                    f'❌ 缺失（期望类型 {col_type}），请执行迁移', 'danger'))
    except Exception as e:
        reports.append(('列检查失败', str(e), 'danger'))

    # SSR 剥离：返回 JSON（原 render_template 的 repair_schema.html 已下线）
    return jsonify({
        'success': True,
        'reports': [{'name': name, 'status': status, 'detail': detail}
                    for name, detail, status in reports],
        'upgrade_output': [],
    })


@login_required
@admin_required
def drawio_diag():
    """drawio 图标库加载诊断（JSON：clibs 与 stencil 探测结果，替代原诊断页）"""
    import os as _os
    import glob
    from urllib.parse import quote
    stencil_dir = os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'static', 'stencils')
    stencil_urls = []
    clibs = ''
    if _os.path.isdir(stencil_dir):
        stencil_urls = [url_for('static', filename='stencils/' + _os.path.basename(f))
                        for f in sorted(glob.glob(_os.path.join(stencil_dir, '*.drawio.xml')))]
        base = request.host_url.rstrip('/')
        clibs = ';'.join('U' + quote(base + u, safe='') for u in stencil_urls)
    return jsonify({'success': True, 'clibs': clibs, 'stencil_urls': stencil_urls})


@deprecated_endpoint('/api/system/ui-version')
@login_required
@admin_required
def system_ui_version():
    """界面版本切换（兼容遗留 POST；保存后回 SPA 系统概览）"""
    from utils.ui_version import set_ui_version
    version = request.form.get('version')
    if version == 'vue':
        set_ui_version('vue')
        current_app.logger.info('用户 [%s] 确认 Vue 单轨界面', current_user.username)
        flash('系统仅使用 Vue 界面', 'success')
    elif version == 'ssr':
        flash('SSR 已移除，系统仅支持 Vue 界面', 'warning')
    return redirect('/app/system/overview')


    # ==================== 侧栏自定义 ====================
@login_required
def system_sidebar():
    """侧栏自定义（GET 已剥离渲染，302 到 SPA；POST 保留 JSON 保存）"""
    from utils.sidebar_config import save_user_sidebar
    if request.method == 'POST':
        payload = request.get_json(silent=True) or {}
        groups_data = payload.get('groups', [])
        if not isinstance(groups_data, list):
            return jsonify({'success': False, 'message': '参数错误'}), 400
        save_user_sidebar(current_user, groups_data)
        return jsonify({'success': True, 'message': '侧栏设置已保存'})
    return redirect('/app/system/sidebar')


@deprecated_endpoint('/api/system/sidebar/reset')
@login_required
def api_sidebar_reset():
    """重置为默认"""
    pref = UserDashboardPreference.query.filter_by(user_id=current_user.id).first()
    if pref:
        pref.sidebar_json = None
        db.session.commit()
    return jsonify({'success': True, 'message': '已重置为系统默认'})


@login_required
def dashboard_reports():
    return redirect('/app/reports')


@login_required
@require_permission('report:view')
def download_template(module):
    """下载批量导入模板 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1890FF', end_color='0969DD', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    templates = {
        'customer': {
            'name': '客户导入模板',
            'headers': ['客户名称', '联系人', '电话', '邮箱', '所属地区', '地市', '地址',
                        '单位类别', '客户等级',
                        '办公室', '有无驻场', '驻场联系人', '驻场联系方式', '驻场办公室',
                        '有无攻防演练', '巡检频率',
                        '来源', '备注'],
        },
        'device': {
            'name': '设备导入模板',
            'headers': ['所属客户', '设备名称', '设备类型', '品牌', '型号', '序列号', 'IP地址', '端口',
                        '登录用户名', '登录密码', '登录方式', '安装位置', '电源配置', '系统版本',
                        '授权开始日期', '授权截止日期', '规则库版本', '是否维修', '是否在用', '备注'],
        },
        'inspection': {
            'name': '巡检记录导入模板',
            'headers': ['客户名称', '标题', '巡检人员', '巡检日期', '巡检地点', '总体状态', '结论', '备注'],
        },
        'fault': {
            'name': '故障记录导入模板',
            'headers': ['客户名称', '标题', '处理人', '故障时间', '故障类型', '故障描述', '故障原因', '解决方案', '处理结果'],
        },
        'spare': {
            'name': '备件导入模板',
            'headers': ['编码', '名称', '分类', '规格', '单位', '最低库存', '备注'],
        },
        'stock': {
            'name': '库存导入模板',
            'headers': ['备件名称', '位置', '数量', '单价'],
        },
    }

    tpl = templates.get(module)
    if not tpl:
        flash('不支持的导入模板类型', 'danger')
        return redirect(url_for('index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tpl['name']

    for col_idx, h in enumerate(tpl['headers'], 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(h) * 2.5, 18)

    import tempfile
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()

    from utils.excel_export import send_temp_export
    return send_temp_export(
        tmp.name, f'{tpl["name"]}_{date.today().isoformat()}.xlsx')
