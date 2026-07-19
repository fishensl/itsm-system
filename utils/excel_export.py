# -*- coding: utf-8 -*-
"""
P1-3.2: 统一 Excel 导出工具

提供：
- export_xlsx(filename, headers, rows): 标准导出，列头 + 行数据
- export_xlsx_with_styles(filename, headers, rows, sheet_name): 增强版（彩色表头）
"""
import os
import tempfile


def export_xlsx(headers, rows, filename, sheet_name='Sheet1', use_styles=True,
                header_color=('1890FF', '096DD9')):
    """生成 Excel 并通过 send_from_directory 发送给客户端

    :param headers: 列名 list，如 ['客户名称', '电话']
    :param rows: 数据 list of list，如 [['张三', '138...'], ...]
    :param filename: 客户端看到的下载文件名
    :param sheet_name: 工作表名
    :param use_styles: 是否使用彩色表头样式
    :param header_color: (起始色, 结束色) 表头渐变，默认蓝；备件用绿 ('52C41A','389E0D')
    :return: (file_path, download_name) — 传给 send_from_directory
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if use_styles:
        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color=header_color[0], end_color=header_color[1], fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
    else:
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val or '')
            if use_styles:
                cell.alignment = Alignment(vertical='center')
                cell.border = thin_border

    # 设置列宽
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(str(h)) * 2.5, 18)

    # 保存到临时文件
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()
    return tmp.name, filename


def cleanup_export_tmp(path):
    """清理导出临时文件"""
    try:
        os.unlink(path)
    except Exception:
        pass
