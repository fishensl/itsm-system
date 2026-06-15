# -*- coding: utf-8 -*-
"""通用工具：导入文件校验、上传安全"""
import os
import re
import tempfile
from flask import flash, redirect, url_for, request
from werkzeug.utils import secure_filename


# 允许的 Excel 扩展名
ALLOWED_EXCEL_EXT = {'.xlsx', '.xls'}
# 允许的图片扩展名
ALLOWED_IMAGE_EXT = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'}
# 导入最大行数（防 DoS）
MAX_IMPORT_ROWS = 5000
# 文件名安全正则：中文/英文/数字/常见分隔符
SAFE_FILENAME_RE = re.compile(r'[\x00-\x1f<>:"/\\|?*]')


def validate_upload(f, allowed_ext, max_size_mb=20):
    """校验上传文件

    返回: (ok: bool, error: str|None, safe_name: str|None)
    """
    if not f or not f.filename:
        return False, '请选择文件', None

    # 1) 扩展名校验
    fname = f.filename
    ext = os.path.splitext(fname)[1].lower()
    if ext not in allowed_ext:
        return False, f'不支持的文件类型: {ext}（允许: {", ".join(sorted(allowed_ext))}）', None

    # 2) 文件名安全处理
    safe_name = secure_filename(fname) or 'upload' + ext
    # 阻止 path traversal
    if '..' in safe_name or safe_name.startswith('.'):
        return False, '文件名不合法', None

    # 3) 大小校验（Flask 的 MAX_CONTENT_LENGTH 也会兜底；此处给友好错误）
    f.seek(0, os.SEEK_END)
    size = f.tell()
    f.seek(0)
    if size > max_size_mb * 1024 * 1024:
        return False, f'文件过大（{size//1024}KB > {max_size_mb}MB）', None

    return True, None, safe_name


def save_temp_upload(f, suffix=None):
    """保存上传文件到临时目录并返回路径"""
    import tempfile
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix or '')
    tmp.close()
    f.save(tmp.name)
    return tmp.name


def open_excel(path, app=None, max_rows=MAX_IMPORT_ROWS):
    """打开上传的 Excel 并校验行数

    返回: (wb, ws, error_response|None)
    error_response 不为 None 时直接返回给调用方
    """
    import openpyxl
    from flask import flash, redirect, url_for
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    except Exception as e:
        if app:
            app.logger.exception('Excel 解析失败: %s', path)
        return None, None, ('无法解析 Excel: %s' % e, 'danger', 'redirect')

    if ws.max_row > max_rows:
        return None, None, (
            '导入文件行数超限（%d > %d）' % (ws.max_row, max_rows),
            'danger',
            'redirect',
        )
    return wb, ws, None


def cleanup_temp_file(path):
    """清理临时文件（吞 OSError）"""
    import os
    try:
        os.unlink(path)
    except Exception:
        pass
